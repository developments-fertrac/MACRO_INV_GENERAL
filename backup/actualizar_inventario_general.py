# actualizar_inventario_integral_optimizado.py
# -*- coding: utf-8 -*-

from __future__ import annotations
import io, re, os, contextlib
from pathlib import Path
from datetime import date, datetime
import pandas as pd
import numpy as np
import msoffcrypto
from unidecode import unidecode
import tempfile
import warnings
import shutil
import win32com

# Limpiar caché corrupto de win32com
try:
    gen_py_path = Path(win32com.__gen_path__)
    if gen_py_path.exists():
        shutil.rmtree(gen_py_path)
        print(f"[INFO] Caché win32com limpiado: {gen_py_path}")
except Exception as e:
    print(f"[ADVERTENCIA] No se pudo limpiar caché win32com: {e}")
# Suprimir advertencias de openpyxl sobre formato condicional
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ==== CONFIG ====
BASE_PATH = Path(__file__).resolve().parent  
# BASE_PATH = Path(r"C:\Users\jperez\Desktop\Tecnologia\Inventario General")
# BASE_PATH = Path(r"C:\MACRO_INVENTARIO_GENERAL")

PASS_INV = "Compras2027"
PASSWORDS_TRY = ["Compras2026", "Compras2027"]

OUTPUT_BASENAME = "$2025 INVENTARIO GENERAL ACTUALIZADO"
APPLY_PASSWORD_TO_OUTPUT = True

# Prefijos para ubicar archivos descargados del ERP
PFX_INV_ACTUALIZADO  = "INVENTARIO GENERAL ACTUALIZADO"
PFX_VAL_GENERAL      = "VALORIZADO GENERAL"
PFX_VAL_FALT_IMPO    = "VALORIZADO FALTANTES IMPO"
PFX_VAL_FALT         = "VALORIZADO FALTANTES"
PFX_VAL_TOBERIN      = "VALORIZADO TOBERIN"
PFX_MARCAS           = "MARCAS"
PFX_DISTRIBUCION     = "DISTRIBUCION DE MATRICES"
PFX_MAYOR_EXISTENCIA = "2025 INVENTARIO MYR EXISTENCIA"
PFX_MATRIZ_USD = "2025 MATRIZ USD"

def buscar_hoja_por_patron(workbook, patron, ignorar_dolares=True):
    """
    Busca una hoja en el workbook que coincida con el patrón dado.

    """
    for sheet_name in workbook.sheetnames:
        nombre_limpio = sheet_name
        if ignorar_dolares:
            # Remover todos los $ del inicio
            nombre_limpio = re.sub(r'^\$+', '', sheet_name).strip()
        
        if patron in nombre_limpio:
            return sheet_name
    return None

def buscar_archivo_por_patron(directorio, patron, ignorar_dolares=True):
    """
    Busca un archivo en el directorio que coincida con el patrón dado.
    """
    dir_path = Path(directorio)
    for archivo in dir_path.glob("*.xlsx"):
        nombre_limpio = archivo.stem  # nombre sin extensión
        if ignorar_dolares:
            # Remover todos los $ del inicio
            nombre_limpio = re.sub(r'^\$+', '', nombre_limpio).strip()
        
        if patron in nombre_limpio:
            return archivo
    return None

# ==== CONFIGURACIÓN DINÁMICA ====

# Para MATRIZ USD
PATRON_MATRIZ_USD = "2025 MATRIZ USD"
PATRON_SHEET_2025 = "2025"

# Para INVENTARIO GENERAL
PATRON_INV_FILE = "2025 INVENTARIO GENERAL"
SHEET_INV_ORIG = "INVENTARIO"
SHEET_INV_COPIA = "INVENTARIO COPIA"
SHEET_INV_LISTA = "INV LISTA PRECIOS"

HEADER_ROW_INV         = 2
HEADER_ROW_INV_LISTA   = 1
HEADER_ROW_VAL         = 9
HEADER_ROW_MATRIZ      = 1
HEADER_ROW_MAYOR_EXIST = 1  

# Columnas a limpiar en INVENTARIO COPIA
COLS_A_LIMPIAR = [
    "REFERENCIA", "NOMBRE LISTA", "NOMBRE ODOO", "NOMBRE MYR",
    "MARCA copia", "INV BODEGA", "EXISTENCIA AGO 26", "COSTO PROMEDIO",
    "LINEA COPIA", "SUB-LINEA COPIA", "LIDER LINEA", "CLASIFICACION",
    "Marca sistema", "Linea sistema", "Sub- linea sistema"
]

# Columnas a traer desde INVENTARIO original
COLS_DESDE_ORIGINAL = ["MARCA copia", "INV BODEGA GERENCIA", "LINEA COPIA", "SUB-LINEA COPIA", "LIDER LINEA", "CLASIFICACION"]

# ==== DEPENDENCIAS (COM) ====
try:
    import win32com.client as win32
    HAS_COM = True
except Exception:
    HAS_COM = False

# ==== UTILS BÁSICAS ====
def log(msg): print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

# Cache para normalización
_NORM_CACHE = {}
def _norm(s: str) -> str:
    if s in _NORM_CACHE:
        return _NORM_CACHE[s]
    t = unidecode(str(s)).lower()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    _NORM_CACHE[s] = t
    return t

def month_abbr_es(dt: date) -> str:
    abrs = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]
    return abrs[dt.month-1]

def exist_col_title_for_today() -> str:
    today = date.today()
    return f"EXISTENCIA {month_abbr_es(today)} {today.day:02d}"

def to_num_str(x):
    """Convierte a referencia numérica segura (string sin .0), manejando separadores de miles."""
    if pd.isna(x): return ""
    
    # Si ya es string, procesar
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return ""
        
        # Si contiene letras, paréntesis o barras, NO es una referencia numérica pura
        # Mantenerlo como está
        if any(c.isalpha() or c in '()/' for c in s):
            return s
        
        # Si solo contiene dígitos, puntos, comas o guiones, es potencialmente numérico
        # Eliminar separadores de miles (punto y coma)
        s_clean = s.replace(".", "").replace(",", "")
        
        # Intentar convertir a número
        try:
            f = float(s_clean)
            if abs(f - int(f)) < 1e-9:
                return str(int(f))
            return str(f)
        except:
            # Si falla la conversión, devolver original
            return s
    
    # Si es numérico (int, float), convertir
    try:
        # Convertir a string primero
        s = str(x).strip()
        
        # Eliminar separadores de miles que puedan venir del formato
        s = s.replace(",", "")
        
        f = float(s)
        if abs(f - int(f)) < 1e-9:
            return str(int(f))
        return str(f)
    except:
        return str(x).strip()
    
def limpiar_referencia(valor):
    """
    Limpia una referencia para eliminar .0 innecesarios y formatear correctamente.
    """
    if valor is None or valor == "":
        return ""
    
    # Convertir a string
    val_str = str(valor).strip()
    
    # Si está vacío después de strip
    if not val_str or val_str in ("None", "nan", "NaN"):
        return ""
    
    # Si termina en .0, quitarlo
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    
    # Si es notación científica, convertir a número normal
    if 'e+' in val_str.lower() or 'E+' in val_str:
        try:
            num = float(val_str)
            if abs(num - int(num)) < 1e-9:
                val_str = str(int(num))
            else:
                val_str = str(num)
        except:
            pass
    
    return val_str

# ==== ARCHIVOS / LECTURA ====
def _strip_dol_tmp(name: str) -> str:
    base = Path(name).stem.replace("~$", "")
    base = re.sub(r"^\$+", "", base)
    return base

def find_by_prefix(basedir: Path, prefix: str, exts=[".xlsx", ".xlsm", ".xls", ".csv"]) -> Path:
    """
    Busca por prefijo normalizado, elige el más reciente.
    NUEVO: Da prioridad a coincidencias exactas antes que parciales.
    """
    pref = _norm(prefix)
    cands = []
    exact_match = None  # NUEVO: almacena coincidencia exacta
    
    for f in basedir.iterdir():
        if not f.is_file() or f.suffix.lower() not in exts:
            continue
        
        nn = _norm(_strip_dol_tmp(f.name))
        
        # NUEVO: Verificar coincidencia exacta primero
        if nn == pref:
            exact_match = f
            log(f"  ✓ Coincidencia EXACTA encontrada: {f.name}")
            break  # Salir inmediatamente si hay coincidencia exacta
        
        # Coincidencias parciales (como antes)
        if nn.startswith(pref) or pref in nn:
            cands.append(f)
            continue
        
        # Búsqueda por tokens (como antes)
        tokens = pref.split()
        if all(t in nn for t in tokens):
            cands.append(f)
    
    # NUEVO: Retornar coincidencia exacta si existe
    if exact_match:
        return exact_match
    
    # Si no hay coincidencia exacta, usar el algoritmo original
    if not cands:
        raise FileNotFoundError(f"No encontré archivos que coincidan con '{prefix}' en {basedir}")
    
    # Ordenar por fecha de modificación más reciente
    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    
    # Log para debugging
    log(f"  → Usando archivo más reciente (sin coincidencia exacta): {cands[0].name}")
    
    return cands[0]


def decrypt_to_stream(xlsx_path: Path, password: str) -> io.BytesIO:
    bio = io.BytesIO()
    with open(xlsx_path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(bio)
    bio.seek(0)
    return bio

def is_encrypted_xlsx(path: Path) -> bool:
    try:
        with open(path, "rb") as f:
            of = msoffcrypto.OfficeFile(f)
            return bool(getattr(of, "is_encrypted", True))
    except Exception:
        return False

def save_bytesio_to_temp(bio: io.BytesIO, stem: str) -> Path:
    tmp = Path(tempfile.gettempdir()) / f"~dec_{stem}_{datetime.now().strftime('%H%M%S')}.xlsx"
    with open(tmp, "wb") as out:
        out.write(bio.getvalue())
    return tmp

def com_convert_to_xlsx(path: Path, passwords: list[str] | None = None) -> Path:
    """Convierte silenciosamente a .xlsx usando COM."""
    passwords = passwords or []
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.Interactive = False
    excel.EnableEvents = False
    excel.ScreenUpdating = False
    
    try: excel.AskToUpdateLinks = False
    except Exception: pass
    try: excel.AutomationSecurity = 3
    except Exception: pass

    encrypted = False
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            with open(path, "rb") as f:
                of = msoffcrypto.OfficeFile(f)
                encrypted = bool(getattr(of, "is_encrypted", False))
        except Exception:
            encrypted = False

    wb = None
    last_err = None
    pw_attempts = (passwords if encrypted else [None] + passwords)

    for pw in pw_attempts:
        try:
            if pw:
                wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True, Password=pw)
            else:
                wb = excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=True, IgnoreReadOnlyRecommended=True)
            break
        except Exception as e:
            last_err = e
            continue

    if wb is None:
        excel.Quit()
        msg = "archivo cifrado sin contraseña válida" if encrypted else "no pude abrir el archivo"
        raise RuntimeError(f"COM no pudo abrir '{path.name}': {msg}. Detalle: {last_err}")

    tmp = Path(tempfile.gettempdir()) / f"~conv_{path.stem}_{datetime.now().strftime('%H%M%S')}.xlsx"
    wb.SaveAs(str(tmp), FileFormat=51)
    wb.Close(SaveChanges=False)
    excel.Quit()
    return tmp

def open_as_excel_source(path: Path, passwords: list[str] | None = None):
    """Devuelve un 'source' para pandas."""
    passwords = passwords or []
    if path.suffix.lower() == ".csv":
        return path
    try:
        with pd.ExcelFile(path, engine="openpyxl"):
            return path
    except Exception as e1:
        err = str(e1).lower()
        if any(k in err for k in ("password", "encrypt", "badzipfile", "not a zip")):
            for pw in passwords:
                try:
                    bio = decrypt_to_stream(path, pw)
                    with pd.ExcelFile(bio, engine="openpyxl"):
                        pass
                    return bio
                except Exception:
                    continue
        if HAS_COM:
            return com_convert_to_xlsx(path, passwords)
        raise

def find_sheet_name_flexible_pd(src, targets=("INVENTARIO","INVENTARIO GENERAL","INV","Sheet1","Sheet 1","Hoja1")) -> str:
    """Elige la mejor hoja."""
    xf = pd.ExcelFile(src, engine="openpyxl")
    names = xf.sheet_names
    if not names:
        raise ValueError("El libro no tiene hojas.")
    norm_map = {_norm(n): n for n in names}
    for t in targets:
        tn = _norm(t)
        if tn in norm_map:
            return norm_map[tn]
    for t in targets:
        tn = _norm(t)
        for kn, real in norm_map.items():
            if tn in kn:
                return real
    return names[0]

def read_excel_header_at(path: Path, sheet: str | int, header_row_visible: int) -> pd.DataFrame:
    """Lee una hoja con header en 'header_row_visible' (1-based)."""
    src = open_as_excel_source(path, PASSWORDS_TRY)
    hdr_idx0 = header_row_visible - 1
    chosen = find_sheet_name_flexible_pd(src, targets=(sheet, "INVENTARIO", "INVENTARIO GENERAL", "INV", "Sheet1", "Sheet 1", "Hoja1")) \
             if isinstance(sheet, str) else sheet
    df = pd.read_excel(src, sheet_name=chosen, engine="openpyxl", header=hdr_idx0)
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")].copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

# ==== LECTURA DE INSUMOS ====
def cargar_inventario_actualizado(base_dir: Path) -> pd.DataFrame:
    """ERP preferido; si no hay, cae en PLANTILLA."""
    try:
        p = find_by_prefix(base_dir, PFX_INV_ACTUALIZADO)
        log(f"Abriendo inventario actualizado (ERP): {p.name}")
        
        # 🔹 AÑADIDO: Manejo de contraseñas
        src = open_as_excel_source(p, PASSWORDS_TRY)
        df = read_excel_header_at(src, sheet="Sheet 1", header_row_visible=1)
        
        idx = {_norm(c): c for c in df.columns}

        ref_col = (
            idx.get("referencia") or idx.get("referencia interna") or idx.get("ref")
            or idx.get("codigo") or idx.get("código")
            or next((real for kn, real in idx.items() if "referenc" in kn or "codigo" in kn or kn.endswith("ref")), None)
        )
        if not ref_col:
            raise KeyError(f"{p.name}: no encuentro columna de Referencia. Encabezados: {list(df.columns)}")

        df = df[~df[ref_col].isna() & (df[ref_col].astype(str).str.strip() != "")].copy()
        
        
        # Buscar específicamente las problemáticas
        refs_problematicas = df[df[ref_col].astype(str).str.contains(r'95\.?276|500\.?845', regex=True, na=False)]
        if len(refs_problematicas) > 0:
            for idx_row, row in refs_problematicas.iterrows():
                val_original = row[ref_col]
        
        # 🔹 NORMALIZACIÓN CRÍTICA: Aplicar to_num_str a cada valor
        df["__REFERENCIA__"] = df[ref_col].apply(to_num_str)    
       

        nom_col     = idx.get("nombre") or "Nombre"
        marca_col   = next((real for kn, real in idx.items()
                            if ("marca/ nombre a mostrar" in kn) or ("marca nombre a mostrar" in kn) or (kn == "marca")), None) \
                      or next((real for kn, real in idx.items() if "marca" in kn and "mostrar" in kn), None)
        linea_col   = next((real for kn, real in idx.items()
                            if ("linea/ nombre a mostrar" in kn) or ("línea/ nombre a mostrar" in kn)), None) \
                      or next((real for kn, real in idx.items() if "linea" in kn and "mostrar" in kn), None)
        sublinea_col = next((real for kn, real in idx.items() if "sub" in kn and "linea" in kn and "mostrar" in kn), None)
        costo_col   = idx.get("costo") or "Costo"

        rename = {}
        if nom_col      in df.columns: rename[nom_col]      = "__NOMBRE__"
        if marca_col    in df.columns: rename[marca_col]    = "__MARCA_SYS__"
        if linea_col    in df.columns: rename[linea_col]    = "__LINEA_SYS__"
        if sublinea_col in df.columns: rename[sublinea_col] = "__SUBLINEA_SYS__"
        if costo_col    in df.columns: rename[costo_col]    = "__COSTO__"
        
        df_final = df.rename(columns=rename)
        

        return df_final
        
    except FileNotFoundError:
        pass

    # 🔹 Usar búsqueda dinámica (ignora $ al inicio)
    p = buscar_archivo_por_patron(base_dir, PATRON_INV_FILE)
    
    if not p:
        # Fallback: intentar con el nombre exacto
        p_pl = base_dir / PATRON_INV_FILE
        if p_pl.exists():
            p = p_pl
        else:
            # Intentar con diferentes variantes
            for pref in ["2025 INVENTARIO GENERAL", "INVENTARIO GENERAL"]:
                try:
                    p = find_by_prefix(base_dir, pref)
                    break
                except Exception:
                    p = None
            
            if p is None:
                raise FileNotFoundError(
                    f"No encontré ni '{PFX_INV_ACTUALIZADO}' ni '{PATRON_INV_FILE}' en {base_dir}"
                )
    log(f"[Fallback] Abriendo plantilla de inventario: {p.name}")
    df = read_excel_header_at(p, sheet=SHEET_INV_ORIG, header_row_visible=HEADER_ROW_INV)
    idx = {_norm(c): c for c in df.columns}

    ref_col = (
        idx.get("referencia") or idx.get("referencia fertrac") or idx.get("referencia interna")
        or idx.get("ref") or idx.get("código") or idx.get("codigo")
        or next((real for kn, real in idx.items() if "referenc" in kn or "codigo" in kn or kn.endswith("ref")), None)
    )
    if not ref_col:
        raise KeyError(f"{p.name}: no encuentro columna 'REFERENCIA'. Encabezados: {list(df.columns)}")

    df = df[~df[ref_col].isna() & (df[ref_col].astype(str).str.strip() != "")].copy()
    df["__REFERENCIA__"] = df[ref_col].apply(to_num_str)

    nombre_odoo = idx.get("nombre odoo") or idx.get("nombre")
    marca_sys   = idx.get("marca sistema")
    linea_sys   = idx.get("linea sistema") or idx.get("línea sistema")
    sub_sys     = idx.get("sub- linea sistema") or idx.get("sub-linea sistema") or idx.get("sub linea sistema")
    costo_prom  = idx.get("costo promedio") or idx.get("costo prom")

    rename = {}
    if nombre_odoo in df.columns: rename[nombre_odoo] = "__NOMBRE__"
    if marca_sys   in df.columns: rename[marca_sys]   = "__MARCA_SYS__"
    if linea_sys   in df.columns: rename[linea_sys]   = "__LINEA_SYS__"
    if sub_sys     in df.columns: rename[sub_sys]     = "__SUBLINEA_SYS__"
    if costo_prom  in df.columns: rename[costo_prom]  = "__COSTO__"

    return df.rename(columns=rename)


# ============================================================================
# FUNCIÓN: ACTUALIZAR REFERENCIAS DESDE FORMATO CODIFICACIÓN
# ============================================================================

def actualizar_referencias_inventario_original(wb, ws_inv_orig, base_path: Path, password: str):
    """
    Actualiza las referencias en la hoja INVENTARIO original basándose en 
    FORMATO CODIFICACIÓN.xlsx antes de empezar el proceso principal.
    
    Args:
        wb: Workbook de Excel abierto
        ws_inv_orig: Worksheet INVENTARIO original
        base_path: Directorio base
        password: Contraseña del archivo
    
    Returns:
        int: Número de referencias actualizadas
    """
    try:
        log("="*70)
        log("ACTUALIZANDO REFERENCIAS DESDE FORMATO CODIFICACIÓN")
        log("="*70)
        
        # 1. BUSCAR ARCHIVO FORMATO CODIFICACIÓN
        archivo_formato = None
        patron_formato = "FORMATO CODIFICACIÓN"
        
        for f in base_path.iterdir():
            if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm'):
                if _norm(patron_formato) in _norm(f.name):
                    archivo_formato = f
                    log(f"Archivo encontrado: {f.name}")
                    break
        
        if not archivo_formato:
            log("⚠️ No se encontró archivo FORMATO CODIFICACIÓN - saltando actualización")
            log("="*70)
            return 0
        
        # 2. CARGAR REFERENCIAS A MODIFICAR
        log("Cargando referencias a modificar...")
        
        try:
            src_formato = open_as_excel_source(archivo_formato, PASSWORDS_TRY)
            
            # Intentar encontrar la hoja
            xf = pd.ExcelFile(src_formato, engine="openpyxl")
            sheet_codificacion = None
            
            for sn in xf.sheet_names:
                sn_norm = _norm(sn)
                if "modificacion" in sn_norm and "ref" in sn_norm:
                    sheet_codificacion = sn
                    log(f"Hoja encontrada: '{sn}'")
                    break
            
            if not sheet_codificacion:
                # Intentar variantes
                for sn in xf.sheet_names:
                    sn_norm = _norm(sn)
                    if "cambio" in sn_norm or "actualizacion" in sn_norm:
                        sheet_codificacion = sn
                        log(f"Hoja encontrada (alternativa): '{sn}'")
                        break
            
            if not sheet_codificacion:
                log("⚠️ No se encontró hoja de modificaciones - usando primera hoja")
                sheet_codificacion = xf.sheet_names[0]
            
            # Leer archivo (encabezados en fila 2, datos desde fila 3)
            df_codificacion = pd.read_excel(
                src_formato,
                sheet_name=sheet_codificacion,
                engine="openpyxl",
                header=1  # Fila 2 (índice 1)
            )
            
            log(f"Total de registros cargados: {len(df_codificacion)}")
            
            # 3. FILTRAR SOLO MODIFICADOS
            # Buscar columna SISTEMA
            col_sistema = None
            for col in df_codificacion.columns:
                if _norm(col) == _norm("SISTEMA"):
                    col_sistema = col
                    break
            
            if not col_sistema:
                log("⚠️ No se encontró columna SISTEMA")
                return 0
            
            # Filtrar
            df_modificaciones = df_codificacion[
                df_codificacion[col_sistema].astype(str).str.strip().str.upper() == 'MODIFICADO'
            ].copy()
            
            log(f"Registros con SISTEMA='MODIFICADO': {len(df_modificaciones)}")
            
            if len(df_modificaciones) == 0:
                log("No hay referencias para modificar")
                log("="*70)
                return 0
            
            # 4. BUSCAR COLUMNAS DE REFERENCIAS
            col_ref_antigua = None
            col_ref_nueva = None
            
            for col in df_modificaciones.columns:
                col_norm = _norm(col)
                if "ref" in col_norm and "fertrac" in col_norm and "modificar" in col_norm:
                    col_ref_antigua = col
                elif "ref" in col_norm and "fertrac" in col_norm and "nueva" in col_norm:
                    col_ref_nueva = col
            
            if not col_ref_antigua or not col_ref_nueva:
                log(f"⚠️ No se encontraron columnas de referencias")
                log(f"   Columnas disponibles: {list(df_modificaciones.columns)}")
                return 0
            
            log(f"Columna antigua: '{col_ref_antigua}'")
            log(f"Columna nueva: '{col_ref_nueva}'")
            
            # 5. CREAR DICCIONARIO DE MAPEO
            mapeo = {}
            for _, row in df_modificaciones.iterrows():
                ref_antigua = str(row[col_ref_antigua]).strip()
                ref_nueva = str(row[col_ref_nueva]).strip()
                
                # Validar que ambas sean válidas
                if (ref_antigua and ref_antigua not in ('nan', 'None', '', 'N/A', '-') and
                    ref_nueva and ref_nueva not in ('nan', 'None', '', 'N/A', '-', 'OK')):
                    
                    # Normalizar ambas referencias
                    ref_antigua_norm = to_num_str(ref_antigua)
                    ref_nueva_norm = to_num_str(ref_nueva)
                    
                    if ref_antigua_norm and ref_nueva_norm:
                        mapeo[ref_antigua_norm] = ref_nueva_norm
            
            log(f"Mapa de reemplazos creado: {len(mapeo)} referencias válidas")
            
            if len(mapeo) == 0:
                log("No hay referencias válidas para reemplazar")
                log("="*70)
                return 0
            
            # Mostrar muestra
            log("")
            log("Muestra de cambios a aplicar:")
            for i, (ref_ant, ref_nue) in enumerate(list(mapeo.items())[:5]):
                log(f"  {ref_ant:40} → {ref_nue}")
            if len(mapeo) > 5:
                log(f"  ... y {len(mapeo) - 5} más")
            
            # 6. ENCONTRAR COLUMNA REFERENCIA EN INVENTARIO ORIGINAL
            log("")
            log("Buscando columna REFERENCIA en INVENTARIO original...")
            
            # ✅ USANDO ws_headers_smart que retorna 3 valores
            hr_orig, hdr_orig, hdrn_orig = ws_headers_smart(
                ws_inv_orig, 
                HEADER_ROW_INV,
                ["REFERENCIA", "REFERENCIA FERTRAC"]
            )
            
            ref_col_idx = hdrn_orig.get(_norm("REFERENCIA")) or \
                         hdrn_orig.get(_norm("REFERENCIA FERTRAC")) or \
                         find_reference_col_idx(hdrn_orig, ws_inv_orig, hr_orig)
            
            log(f"Columna REFERENCIA encontrada: índice {ref_col_idx}")
            
            # 7. DETERMINAR ÚLTIMA FILA
            last_row_orig = ws_last_row(ws_inv_orig, ref_col_idx, HEADER_ROW_INV)
            
            # Ajustar por pivots
            pivot_top = ws_first_pivot_row(ws_inv_orig)
            if pivot_top and pivot_top > HEADER_ROW_INV:
                last_row_orig = min(last_row_orig, pivot_top - 1)
            
            log(f"Rango a procesar: filas {HEADER_ROW_INV + 1} a {last_row_orig}")
            
            # 8. LEER TODAS LAS REFERENCIAS
            log("")
            log("Leyendo referencias actuales...")
            
            referencias_actuales = read_range_as_array(
                ws_inv_orig, 
                HEADER_ROW_INV + 1, 
                last_row_orig, 
                ref_col_idx
            )
            
            # Normalizar
            referencias_norm = [to_num_str(r) for r in referencias_actuales]
            
            log(f"Total de referencias leídas: {len(referencias_norm)}")
            
            # 9. REALIZAR REEMPLAZOS
            log("")
            log("Aplicando reemplazos...")
            
            nuevas_referencias = []
            reemplazos_realizados = 0
            referencias_actualizadas = set()
            
            # 🆕 REGISTRO DETALLADO PARA REPORTE
            cambios_exitosos = []
            
            for i, ref_norm in enumerate(referencias_norm):
                if ref_norm in mapeo:
                    ref_nueva = mapeo[ref_norm]
                    nuevas_referencias.append(ref_nueva)
                    reemplazos_realizados += 1
                    referencias_actualizadas.add(ref_norm)
                    
                    # 🆕 Guardar para reporte
                    cambios_exitosos.append({
                        'FILA_EXCEL': HEADER_ROW_INV + 1 + i,
                        'REFERENCIA_ANTIGUA': referencias_actuales[i],  # Original sin normalizar
                        'REFERENCIA_NUEVA': ref_nueva,
                        'ESTADO': 'ACTUALIZADO'
                    })
                    
                    if reemplazos_realizados <= 5:
                        log(f"  Fila {HEADER_ROW_INV + 1 + i}: {ref_norm:40} → {ref_nueva}")
                    elif reemplazos_realizados == 6:
                        log(f"  ... procesando más reemplazos ...")
                else:
                    nuevas_referencias.append(referencias_actuales[i])
            
            log(f"Reemplazos realizados: {reemplazos_realizados}")
            
            # 10. ESCRIBIR DE VUELTA
            if reemplazos_realizados > 0:
                log("")
                log("Escribiendo referencias actualizadas...")
                
                write_range_as_array(
                    ws_inv_orig,
                    HEADER_ROW_INV + 1,
                    ref_col_idx,
                    nuevas_referencias
                )
                
                log(f"✅ Referencias actualizadas en INVENTARIO original")
            else:
                log("No se realizaron cambios")
            
            # 11. VERIFICAR REFERENCIAS NO ENCONTRADAS
            refs_no_encontradas = set(mapeo.keys()) - referencias_actualizadas
            
            # 🆕 REGISTRO DE NO ENCONTRADAS PARA REPORTE
            cambios_no_encontrados = []
            for ref_antigua in refs_no_encontradas:
                ref_nueva = mapeo[ref_antigua]
                cambios_no_encontrados.append({
                    'FILA_EXCEL': 'N/A',
                    'REFERENCIA_ANTIGUA': ref_antigua,
                    'REFERENCIA_NUEVA': ref_nueva,
                    'ESTADO': 'NO ENCONTRADA EN INVENTARIO'
                })
            
            if refs_no_encontradas:
                log("")
                log(f"⚠️ Referencias NO encontradas en inventario: {len(refs_no_encontradas)}")
                for ref in list(refs_no_encontradas)[:5]:
                    log(f"    - {ref}")
                if len(refs_no_encontradas) > 5:
                    log(f"    ... y {len(refs_no_encontradas) - 5} más")
            
            # 🆕 12. GENERAR REPORTE EXCEL
            log("")
            log("Generando reporte de cambios...")
            
            try:
                # Crear DataFrame con TODOS los cambios
                df_reporte = pd.DataFrame(cambios_exitosos + cambios_no_encontrados)
                
                # Ordenar: primero exitosos, luego no encontrados
                df_reporte['_ORDEN'] = df_reporte['ESTADO'].apply(
                    lambda x: 0 if x == 'ACTUALIZADO' else 1
                )
                df_reporte = df_reporte.sort_values(['_ORDEN', 'REFERENCIA_ANTIGUA'])
                df_reporte = df_reporte.drop(columns=['_ORDEN'])
                
                # Generar nombre del archivo
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_reporte = f"REPORTE_CAMBIOS_REFERENCIAS_{timestamp}.xlsx"
                ruta_reporte = base_path / nombre_reporte
                
                # Guardar usando openpyxl para mejor formato
                with pd.ExcelWriter(ruta_reporte, engine='openpyxl') as writer:
                    # Hoja 1: Cambios exitosos
                    if cambios_exitosos:
                        df_exitosos = pd.DataFrame(cambios_exitosos)
                        df_exitosos.to_excel(writer, sheet_name='REFERENCIAS ACTUALIZADAS', index=False)
                    
                    # Hoja 2: No encontradas
                    if cambios_no_encontrados:
                        df_no_encontrados = pd.DataFrame(cambios_no_encontrados)
                        df_no_encontrados.to_excel(writer, sheet_name='NO ENCONTRADAS', index=False)
                    
                    # Hoja 3: Resumen
                    resumen_data = {
                        'MÉTRICA': [
                            'Total referencias a cambiar',
                            'Referencias actualizadas',
                            'Referencias NO encontradas',
                            'Fecha de proceso',
                            'Archivo origen'
                        ],
                        'VALOR': [
                            len(mapeo),
                            reemplazos_realizados,
                            len(refs_no_encontradas),
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            archivo_formato.name
                        ]
                    }
                    df_resumen = pd.DataFrame(resumen_data)
                    df_resumen.to_excel(writer, sheet_name='RESUMEN', index=False)
                
                # Ajustar anchos de columna
                try:
                    from openpyxl import load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    wb_reporte = load_workbook(ruta_reporte)
                    
                    for sheet_name in wb_reporte.sheetnames:
                        ws = wb_reporte[sheet_name]
                        
                        # Ajustar anchos
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        # Formatear encabezados
                        if ws.max_row > 0:
                            for cell in ws[1]:
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Colorear estados
                        if sheet_name in ['REFERENCIAS ACTUALIZADAS', 'NO ENCONTRADAS']:
                            estado_col = None
                            for idx, cell in enumerate(ws[1], start=1):
                                if cell.value == 'ESTADO':
                                    estado_col = idx
                                    break
                            
                            if estado_col:
                                for row in range(2, ws.max_row + 1):
                                    cell = ws.cell(row=row, column=estado_col)
                                    if cell.value == 'ACTUALIZADO':
                                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                        cell.font = Font(color="006100")
                                    elif cell.value == 'NO ENCONTRADA EN INVENTARIO':
                                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                        cell.font = Font(color="9C0006")
                    
                    wb_reporte.save(ruta_reporte)
                    
                except Exception as e_formato:
                    log(f"  ⚠️ No se pudo aplicar formato al reporte: {e_formato}")
                
                log(f"✅ Reporte generado: {nombre_reporte}")
                log(f"   📁 Ubicación: {ruta_reporte}")
                log(f"   📊 Hojas:")
                log(f"      - REFERENCIAS ACTUALIZADAS: {len(cambios_exitosos)} registros")
                log(f"      - NO ENCONTRADAS: {len(cambios_no_encontrados)} registros")
                log(f"      - RESUMEN: Estadísticas del proceso")
                
            except Exception as e_reporte:
                log(f"⚠️ Error al generar reporte: {e_reporte}")
                import traceback
                log(traceback.format_exc())
            
            log("="*70)
            log("")
            
            return reemplazos_realizados
            
        except Exception as e:
            log(f"⚠️ Error al procesar FORMATO CODIFICACIÓN: {e}")
            import traceback
            log(traceback.format_exc())
            return 0
        
    except Exception as e:
        log(f"⚠️ Error en actualización de referencias: {e}")
        import traceback
        log(traceback.format_exc())
        return 0
    
def cargar_valorizado(base_dir: Path, prefix: str) -> pd.DataFrame:
    """Lee VALORIZADO* (header visible en fila 9)."""
    p = find_by_prefix(base_dir, prefix)
    log(f"Abrir: {p.name}")
    src = open_as_excel_source(p, PASSWORDS_TRY)

    if p.suffix.lower() == ".csv":
        df_all = pd.read_csv(src, header=None, dtype=str)
    else:
        df_all = pd.read_excel(src, sheet_name=0, engine="openpyxl", header=None)

    hdr_row0 = HEADER_ROW_VAL - 1
    if hdr_row0 >= len(df_all):
        raise ValueError(f"{p.name}: HEADER_ROW_VAL={HEADER_ROW_VAL} supera el número de filas.")

    df = df_all.iloc[hdr_row0:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.iloc[0]]
    df = df.iloc[1:].reset_index(drop=True)
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]

    idx = {_norm(c): c for c in df.columns}

    # FORZAR a buscar EXACTAMENTE "Referencia interna"
    refc = idx.get("referencia interna")

    # Si no encuentra, buscar alternativas PERO mostrar advertencia
    if not refc:
        log(f"  ⚠️ ADVERTENCIA: No se encontró columna 'Referencia interna' en {p.name}")
        log(f"     Columnas disponibles: {list(df.columns)}")
        refc = idx.get("referencia") or idx.get("ref") \
            or next((real for kn, real in idx.items() if "referenc" in kn), None)
        if refc:
            log(f"     Usando columna alternativa: '{refc}'")

    cant = idx.get("cantidad")
    if not cant:
        log(f"  ⚠️ ADVERTENCIA: No se encontró columna 'Cantidad' en {p.name}")
        log(f"     Columnas disponibles: {list(df.columns)}")
        cant = next((real for kn, real in idx.items() if kn.startswith("cant")), None)
        if cant:
            log(f"     Usando columna alternativa: '{cant}'")

    if not refc: raise KeyError(f"{p.name}: no encuentro 'Referencia interna'. Encabezados: {list(df.columns)}")
    if not cant: raise KeyError(f"{p.name}: no encuentro 'Cantidad'. Encabezados: {list(df.columns)}")

    out = pd.DataFrame()
    # 🔹 NORMALIZACIÓN: Aplicar to_num_str
    out["__REF_INT__"] = df[refc].apply(to_num_str)
    out["__CANT__"]    = pd.to_numeric(df[cant], errors="coerce").fillna(0.0)
    
    # 🔹 VERIFICACIÓN: Buscar referencias problemáticas
    for ref_buscar in ["95276", "500845"]:
        if ref_buscar in out["__REF_INT__"].values:
            cantidad = out[out["__REF_INT__"] == ref_buscar]["__CANT__"].iloc[0]
    
    return out

def cargar_valorizado_desde_ruta(archivo_path: Path) -> pd.DataFrame:
    """Lee VALORIZADO desde ruta específica."""
    log(f"Abriendo: {archivo_path.name}")
    src = open_as_excel_source(archivo_path, PASSWORDS_TRY)

    if archivo_path.suffix.lower() == ".csv":
        df_all = pd.read_csv(src, header=None, dtype=str)
    else:
        df_all = pd.read_excel(src, sheet_name=0, engine="openpyxl", header=None)

    hdr_row0 = HEADER_ROW_VAL - 1
    df = df_all.iloc[hdr_row0:].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.iloc[0]]
    df = df.iloc[1:].reset_index(drop=True)
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]

    idx = {_norm(c): c for c in df.columns}
    refc = idx.get("referencia interna")
    cant = idx.get("cantidad")

    if not refc: raise KeyError(f"{archivo_path.name}: no encuentro 'Referencia interna'")
    if not cant: raise KeyError(f"{archivo_path.name}: no encuentro 'Cantidad'")

    out = pd.DataFrame()
    out["__REF_INT__"] = df[refc].apply(to_num_str)
    out["__CANT__"]    = pd.to_numeric(df[cant], errors="coerce").fillna(0.0)
    
    return out

def cargar_matriz_usd(base_dir: Path) -> pd.DataFrame:
    """
    Carga el archivo MATRIZ USD, hoja 2025.
    """
    try:
        # 🔹 Usar búsqueda dinámica (ignora $ al inicio)
        p = buscar_archivo_por_patron(base_dir, PATRON_MATRIZ_USD)
        
        if not p:
            log(f"⚠ No se encontró con búsqueda dinámica, intentando método tradicional...")
            p = find_by_prefix(base_dir, PFX_MATRIZ_USD)
        
        log(f"Abriendo Matriz USD: {p.name}")
        
        src = open_as_excel_source(p, PASSWORDS_TRY)
        
        xf = pd.ExcelFile(src, engine="openpyxl")
        
        # 🔹 Búsqueda dinámica de hoja (ignora $ al inicio)
        sheet_found = None
        for sn in xf.sheet_names:
            nombre_limpio = re.sub(r'^\$+', '', sn).strip()
            if PATRON_SHEET_2025 in nombre_limpio or _norm(nombre_limpio) == _norm(PATRON_SHEET_2025):
                sheet_found = sn
                log(f"  ✓ Hoja encontrada: '{sn}'")
                break
        
        if not sheet_found:
            sheet_found = xf.sheet_names[0]
            log(f"  ⚠ No se encontró hoja con '2025', usando: '{sheet_found}'")
        
        df_raw = pd.read_excel(src, sheet_name=sheet_found, engine="openpyxl", header=None)
        
        header_row_idx = None
        for idx in range(min(20, len(df_raw))):
            row_values = df_raw.iloc[idx].astype(str).str.lower()
            has_ref = any("referencia" in str(v).lower() and "fertrac" in str(v).lower() for v in df_raw.iloc[idx])
            has_desc = any("descripcion" in str(v).lower() and "lista" in str(v).lower() for v in df_raw.iloc[idx])
            
            if has_ref or has_desc:
                header_row_idx = idx
                break
        
        if header_row_idx is None:
            max_non_empty = 0
            for idx in range(min(10, len(df_raw))):
                non_empty = df_raw.iloc[idx].notna().sum()
                if non_empty > max_non_empty:
                    max_non_empty = non_empty
                    header_row_idx = idx
            log(f"  Usando fila {header_row_idx + 1} como encabezado (más valores no vacíos)")
        
        df = pd.read_excel(src, sheet_name=sheet_found, engine="openpyxl", header=header_row_idx)
        
        df.columns = [str(c).strip() if not str(c).startswith("Unnamed") and str(c) != "nan" else f"_COL_{i}" 
                      for i, c in enumerate(df.columns)]        
      
        idx = {_norm(c): c for c in df.columns}
        
        #BUSCAR: REFERENCIA INVENTARIO FERTRAC
        ref_col = None
        for col_name in df.columns:
            col_norm = _norm(col_name)
            if "referencia" in col_norm and ("fertrac" in col_norm or "inventario" in col_norm):
                ref_col = col_name
                break
        
        if not ref_col:
            for col_name in df.columns[:5]:
                non_null = df[col_name].notna().sum()
                if non_null > 10:
                    sample = df[col_name].dropna().astype(str).head(5)
                    if any("FP-" in str(v) or str(v).replace("-", "").isdigit() for v in sample):
                        ref_col = col_name
                        log(f"  Usando columna '{col_name}' como REFERENCIA (detectada por patrón)")
                        break
        
        #BUSCAR: DESCRIPCION LISTA PRECIOS
        desc_col = None
        for col_name in df.columns:
            col_norm = _norm(col_name)
            if "descripcion" in col_norm and "lista" in col_norm and "precio" in col_norm:
                desc_col = col_name
                break
        
        if not desc_col:
            for col_name in df.columns:
                if col_name == ref_col:
                    continue
                non_null = df[col_name].notna().sum()
                if non_null > 10:
                    sample = df[col_name].dropna().astype(str).head(5)
                    avg_len = sum(len(str(v)) for v in sample) / len(sample) if len(sample) > 0 else 0
                    if avg_len > 15:
                        desc_col = col_name
                        log(f"  Usando columna '{col_name}' como DESCRIPCION (detectada por longitud)")
                        break
        
        # BUSCAR: REFERENCIA LISTA DE PRECIOS
        ref_lista_col = None
        for col_name in df.columns:
            col_norm = _norm(col_name)
            # Buscar variantes del nombre
            if ("referencia" in col_norm and "lista" in col_norm and "precio" in col_norm):
                ref_lista_col = col_name
                break
        
        # Si no se encuentra por nombre exacto, buscar alternativas
        if not ref_lista_col:
            for col_name in df.columns:
                col_norm = _norm(col_name)
                if col_name == ref_col or col_name == desc_col:
                    continue
                # Buscar "REF LISTA", "CODIGO LISTA", etc.
                if ("ref" in col_norm or "codigo" in col_norm) and "lista" in col_norm:
                    ref_lista_col = col_name
                    log(f"Columna REFERENCIA LISTA encontrada (alternativa): '{col_name}'")
                    break
        
        if not ref_col:
            raise KeyError(f"No encontré columna 'REFERENCIA INVENTARIO FERTRAC' en {p.name}. Columnas: {list(df.columns)}")
        if not desc_col:
            raise KeyError(f"No encontré columna 'DESCRIPCION LISTA PRECIOS' en {p.name}. Columnas: {list(df.columns)}")
        
        #ADVERTENCIA si no se encuentra REFERENCIA LISTA
        if not ref_lista_col:
            log(f"  ⚠ ADVERTENCIA: No se encontró columna 'REFERENCIA LISTA DE PRECIOS' en {p.name}")
            log(f"     Columnas disponibles: {list(df.columns)}")
        
        df = df[~df[ref_col].isna() & (df[ref_col].astype(str).str.strip() != "")].copy()
        
        out = pd.DataFrame()
        out["__REF_MATRIZ__"] = df[ref_col].apply(to_num_str)
        out["__DESC_LISTA__"] = df[desc_col].fillna("")
        
        #Agregar REFERENCIA LISTA DE PRECIOS
        if ref_lista_col:
            out["__REF_LISTA_PRECIOS__"] = df[ref_lista_col].apply(to_num_str)
        else:
            out["__REF_LISTA_PRECIOS__"] = ""  # Columna vacía si no se encuentra
        
        out = out.drop_duplicates(subset=["__REF_MATRIZ__"], keep="first")
        
        if ref_lista_col:
            no_vacias = out["__REF_LISTA_PRECIOS__"].astype(str).str.strip().ne("").sum()
        
        return out
        
    except FileNotFoundError:
        log(f"⚠ ADVERTENCIA: No se encontró el archivo '{PFX_MATRIZ_USD}'.")
        return pd.DataFrame(columns=["__REF_MATRIZ__", "__DESC_LISTA__", "__REF_LISTA_PRECIOS__"])
    except Exception as e:
        log(f"⚠ ERROR al cargar Matriz USD: {e}")
        import traceback
        log(traceback.format_exc())
        return pd.DataFrame(columns=["__REF_MATRIZ__", "__DESC_LISTA__", "__REF_LISTA_PRECIOS__"])
    
def cargar_marcas(base_dir: Path) -> set:
    """
    Carga el archivo MARCAS y retorna un set con las marcas propias.
    """
    try:
        p = find_by_prefix(base_dir, PFX_MARCAS)
        log(f"Abriendo archivo MARCAS: {p.name}")
        
        src = open_as_excel_source(p, PASSWORDS_TRY)        
        df = pd.read_excel(src, sheet_name=0, engine="openpyxl", header=None)      
        marcas_propias = set()
        
        for col_idx in range(min(3, len(df.columns))):
            for val in df[col_idx].dropna():
                val_str = str(val).strip().upper()
                if val_str and val_str not in ("", "NONE", "NAN", "MARCA", "MARCAS"):
                    if any(c.isalpha() for c in val_str):
                        marcas_propias.add(val_str)
        
        log(f"{len(marcas_propias)} marcas propias cargadas")
        return marcas_propias
        
    except FileNotFoundError:
        log(f"⚠ ADVERTENCIA: No se encontró el archivo '{PFX_MARCAS}'")
        return set()
    except Exception as e:
        log(f"⚠ ERROR al cargar MARCAS: {e}")
        import traceback
        log(traceback.format_exc())
        return set()


def cargar_distribucion(base_dir: Path) -> dict:
    """
    Carga el archivo DISTRIBUCIÓN DE MATRICES.
    """
    try:
        p = find_by_prefix(base_dir, PFX_DISTRIBUCION)
        log(f"Abriendo archivo DISTRIBUCIÓN: {p.name}")
        
        src = open_as_excel_source(p, PASSWORDS_TRY)
        xf = pd.ExcelFile(src, engine="openpyxl")
        sheet_name = None
        for sn in xf.sheet_names:
            if "DISTRIBUCION" in _norm(sn) or "MATRICES" in _norm(sn):
                sheet_name = sn
                break
        
        if not sheet_name:
            sheet_name = xf.sheet_names[0]
        
        df_raw = pd.read_excel(src, sheet_name=sheet_name, engine="openpyxl", header=None)
        header_row = None
        for idx in range(min(10, len(df_raw))):
            row_str = ' '.join([str(v).upper() for v in df_raw.iloc[idx] if pd.notna(v)])
            if "LINEA" in row_str and "GESTOR" in row_str:
                header_row = idx
                break
        
        if header_row is None:
            header_row = 2         
        df = pd.read_excel(src, sheet_name=sheet_name, engine="openpyxl", header=header_row)
        df.columns = [str(c).strip() for c in df.columns]
        
        idx = {_norm(c): c for c in df.columns}
        linea_col = (
            idx.get("linea") or idx.get("línea") or idx.get("marca")
            or next((real for kn, real in idx.items() if "linea" in kn or "línea" in kn), None)
        )

        gestor_col = (
            idx.get("gestor") or idx.get("lider") or idx.get("líder") 
            or next((real for kn, real in idx.items() if "gestor" in kn or "lider" in kn), None)
        )
        
        clasif_col = (
            idx.get("categoria") or idx.get("categoría") or idx.get("clasificacion") 
            or idx.get("clasificación") or idx.get("tipo")
            or next((real for kn, real in idx.items() if "categ" in kn or "clasificac" in kn), None)
        )
        
        if not linea_col:
            log(f"  ⚠ No se encontró columna de LINEA/MARCA")
            return {'gestor': {}, 'clasificacion': {}}
        
        gestor_map = {}
        clasif_map = {}

        for idx_row, row in df.iterrows():
            linea_val = row[linea_col] if linea_col in row.index else None
            if pd.isna(linea_val) or str(linea_val).strip() == "":
                continue
            
            linea_str = str(linea_val).strip().upper()
            
            linea_key = re.sub(r'\s*\([^)]*\)\s*', '', linea_str).strip()
            
            if not linea_key:
                continue
        
            if gestor_col and gestor_col in row.index:
                gestor_val = row[gestor_col]
                if not pd.isna(gestor_val) and str(gestor_val).strip():
                    gestor_map[linea_key] = str(gestor_val).strip()
            
            if clasif_col and clasif_col in row.index:
                clasif_val = row[clasif_col]
                if not pd.isna(clasif_val) and str(clasif_val).strip():
                    clasif_map[linea_key] = str(clasif_val).strip()
        
        log(f"{len(gestor_map)} gestores cargados")
        log(f"{len(clasif_map)} clasificaciones cargadas")
        
        return {
            'gestor': gestor_map,
            'clasificacion': clasif_map
        }
        
    except FileNotFoundError:
        log(f"⚠ ADVERTENCIA: No se encontró el archivo '{PFX_DISTRIBUCION}'")
        return {'gestor': {}, 'clasificacion': {}}
    except Exception as e:
        log(f"⚠ ERROR al cargar DISTRIBUCIÓN: {e}")
        import traceback
        log(traceback.format_exc())
        return {'gestor': {}, 'clasificacion': {}}
    
def cargar_mayor_existencia(base_dir: Path) -> pd.DataFrame:
    """
    Carga el archivo MAYOR EXISTENCIA, hoja COSTOS INV FINAL.
    Retorna REFERENCIA FERTRAC y REM EN CONSIG
    """
    try:
        p = find_by_prefix(base_dir, PFX_MAYOR_EXISTENCIA)
        log(f"Abriendo Mayor Existencia: {p.name}")
        
        src = open_as_excel_source(p, PASSWORDS_TRY)
        
        # Buscar hoja COSTOS INV FINAL
        xf = pd.ExcelFile(src, engine="openpyxl")
        sheet_found = None
        
        for sn in xf.sheet_names:
            sn_norm = _norm(sn)
            if "costos" in sn_norm and "inv" in sn_norm and "final" in sn_norm:
                sheet_found = sn
                log(f"Hoja encontrada: '{sn}'")
                break
        
        if not sheet_found:
            # Buscar alternativas
            for sn in xf.sheet_names:
                sn_norm = _norm(sn)
                if "costo" in sn_norm or "final" in sn_norm:
                    sheet_found = sn
                    log(f"Hoja encontrada (alternativa): '{sn}'")
                    break
        
        if not sheet_found:
            sheet_found = xf.sheet_names[0]
            log(f"  ⚠ Usando primera hoja: '{sheet_found}'")
        
        # Leer archivo buscando el encabezado
        df_raw = pd.read_excel(src, sheet_name=sheet_found, engine="openpyxl", header=None)
        
        # Buscar fila de encabezado
        header_row_idx = None
        for idx in range(min(20, len(df_raw))):
            row_str = ' '.join([str(v).upper() for v in df_raw.iloc[idx] if pd.notna(v)])
            if ("REFERENCIA" in row_str or "REF" in row_str) and ("CONSIG" in row_str or "REM" in row_str):
                header_row_idx = idx
                log(f"Encabezado encontrado en fila {idx + 1}")
                break
        
        if header_row_idx is None:
            # Usar header_row configurado
            header_row_idx = HEADER_ROW_MAYOR_EXIST - 1
            log(f"  ⚠ Usando fila de encabezado configurada: {HEADER_ROW_MAYOR_EXIST}")
        
        # Leer con el encabezado correcto
        df = pd.read_excel(src, sheet_name=sheet_found, engine="openpyxl", header=header_row_idx)
        
        df.columns = [str(c).strip() for c in df.columns]
        
        idx = {_norm(c): c for c in df.columns}
        
        #BUSCAR: REFERENCIA FERTRAC
        ref_col = None
        for col_name in df.columns:
            col_norm = _norm(col_name)
            if "referencia" in col_norm and "fertrac" in col_norm:
                ref_col = col_name
                log(f"Columna REFERENCIA FERTRAC encontrada: '{col_name}'")
                break
        
        if not ref_col:
            # Buscar solo "REFERENCIA"
            for col_name in df.columns:
                col_norm = _norm(col_name)
                if col_norm == "referencia" or col_norm.startswith("referencia "):
                    ref_col = col_name
                    log(f"Columna REFERENCIA encontrada: '{col_name}'")
                    break
        
        if not ref_col:
            # Buscar simplemente "REF" o columnas que contengan "referenc"
            for col_name in df.columns[:10]:
                col_norm = _norm(col_name)
                if "ref" in col_norm or "codigo" in col_norm:
                    ref_col = col_name
                    log(f"Columna REFERENCIA encontrada (alternativa): '{col_name}'")
                    break
        
        #BUSCAR: REM EN CONSIG (columna AI)
        rem_consig_col = None
        
        # Buscar exactamente "REM EN CONSIG"
        for col_name in df.columns:
            col_norm = _norm(col_name)
            if col_norm == "rem en consig":
                rem_consig_col = col_name
                log(f"Columna REM EN CONSIG encontrada: '{col_name}'")
                break
        
        if not rem_consig_col:
            # Buscar variantes
            for col_name in df.columns:
                col_norm = _norm(col_name)
                if "rem" in col_norm and "consig" in col_norm:
                    rem_consig_col = col_name
                    log(f"Columna REM EN CONSIG encontrada (variante): '{col_name}'")
                    break
        
        if not rem_consig_col:
            # Buscar por posición (columna AI = índice 34 en Excel, pero en pandas puede variar)
            # Buscar columnas que contengan "rem" o "consig"
            for col_name in df.columns:
                col_norm = _norm(col_name)
                if "consignacion" in col_norm or "consig" in col_norm:
                    rem_consig_col = col_name
                    log(f"  ⚠ Usando columna que contiene 'consig': '{col_name}'")
                    break
        
        if not ref_col:
            raise KeyError(f"No encontré columna REFERENCIA en {p.name}. Columnas: {list(df.columns)}")
        if not rem_consig_col:
            raise KeyError(f"No encontré columna REM EN CONSIG en {p.name}. Columnas: {list(df.columns)}")
        
        # Filtrar filas válidas
        df = df[~df[ref_col].isna() & (df[ref_col].astype(str).str.strip() != "")].copy()
        
        # Construir DataFrame de salida
        out = pd.DataFrame()
        out["__REF_MAYOR__"] = df[ref_col].apply(to_num_str)
        out["__REM_CONSIG__"] = pd.to_numeric(df[rem_consig_col], errors="coerce").fillna(0)
        
        # Eliminar duplicados
        out = out.drop_duplicates(subset=["__REF_MAYOR__"], keep="first")
        
        # Estadísticas
        valores_no_cero = (out["__REM_CONSIG__"] != 0).sum()
        log(f"Mayor Existencia cargada: {len(out)} referencias")
        log(f"REM EN CONSIG: {valores_no_cero} valores diferentes de cero")
        
        return out
        
    except FileNotFoundError:
        log(f"⚠ ADVERTENCIA: No se encontró el archivo '{PFX_MAYOR_EXISTENCIA}'.")
        return pd.DataFrame(columns=["__REF_MAYOR__", "__REM_CONSIG__"])
    except Exception as e:
        log(f"⚠ ERROR al cargar Mayor Existencia: {e}")
        import traceback
        log(traceback.format_exc())
        return pd.DataFrame(columns=["__REF_MAYOR__", "__REM_CONSIG__"])

def aplicar_reglas_marcas_propias(ws_inv_copia, start_data_row: int, last_row: int, 
                                   ref_col_idx: int, hdrn_copia: dict, 
                                   marcas_propias: set, distribucion: dict):
    """
    Aplica las reglas de negocio para referencias de marcas propias:
    """
    try:
        log("Aplicando reglas para marcas propias...")
        
        col_linea_copia = hdrn_copia.get(_norm("LINEA COPIA"))
        col_marca_sistema = hdrn_copia.get(_norm("MARCA SISTEMA")) or hdrn_copia.get(_norm("Marca sistema"))
        col_marca_copia = hdrn_copia.get(_norm("MARCA COPIA")) or hdrn_copia.get(_norm("MARCA copia"))
        col_inv_bodega_ger = hdrn_copia.get(_norm("INV BODEGA GERENCIA"))
        col_sublinea_copia = hdrn_copia.get(_norm("SUB-LINEA COPIA"))
        col_lider_linea = hdrn_copia.get(_norm("LIDER LINEA"))
        col_clasificacion = hdrn_copia.get(_norm("CLASIFICACION"))
        
        if not col_linea_copia or not col_marca_sistema:
            log("  ⚠ No se encontraron columnas necesarias (LINEA COPIA o MARCA SISTEMA)")
            log(f"    LINEA COPIA: {'✓' if col_linea_copia else '✗'}")
            log(f"    MARCA SISTEMA: {'✓' if col_marca_sistema else '✗'}")
            return last_row
        
        # FASE 1: ELIMINAR REFERENCIAS TIPO "0041R"
        log(" Identificando referencias tipo '0041R' para eliminar...")
        cols_to_read = [ref_col_idx]
        data = read_multiple_columns_optimized(ws_inv_copia, start_data_row, last_row, cols_to_read)
        referencias = data[ref_col_idx]
        
        filas_a_eliminar = []
        for i in range(len(referencias)):
            ref = str(referencias[i]).strip() if referencias[i] not in (None, "", "None") else ""
            if ref and ref.upper() == '0041R':
                filas_a_eliminar.append((i, ref))
                    
        if filas_a_eliminar:
            log(f"  Eliminando {len(filas_a_eliminar)} referencias tipo '0041R':")
            for idx, ref in filas_a_eliminar[:5]:
                log(f"    - {ref}")
            if len(filas_a_eliminar) > 5:
                log(f"    ... y {len(filas_a_eliminar) - 5} más")
            
            # Eliminar en orden inverso
            for idx, ref in sorted(filas_a_eliminar, reverse=True):
                fila_excel = start_data_row + idx
                try:
                    ws_inv_copia.Rows(fila_excel).Delete()
                except Exception as e:
                    log(f"    ⚠ Error al eliminar fila {fila_excel} ({ref}): {e}")
            
            # Actualizar last_row
            last_row = last_row - len(filas_a_eliminar)
            log(f"  {len(filas_a_eliminar)} filas eliminadas. Nuevo rango: hasta fila {last_row}")
        else:
            log("  No se encontraron referencias tipo '####L' para eliminar")
        
        # FASE 2: APLICAR FILTROS Y ACTUALIZAR COLUMNAS
        log("  Fase 2: Aplicando filtros y actualizando campos...")
        
        # Volver a leer los datos DESPUÉS de eliminar filas
        cols_to_read = [ref_col_idx, col_linea_copia, col_marca_sistema]
        data = read_multiple_columns_optimized(ws_inv_copia, start_data_row, last_row, cols_to_read)
        
        referencias = data[ref_col_idx]
        lineas_copia = data[col_linea_copia]
        marcas_sistema = data[col_marca_sistema]
        
        filas_a_procesar = []
        
        log(f"  Analizando {len(referencias)} registros con filtros...")
        
        for i in range(len(referencias)):
            ref = str(referencias[i]).strip() if referencias[i] not in (None, "", "None") else ""
            linea = str(lineas_copia[i]).strip() if lineas_copia[i] not in (None, "", "None") else ""
            marca = str(marcas_sistema[i]).strip() if marcas_sistema[i] not in (None, "", "None") else ""
            
            linea_upper = linea.upper()
            marca_upper = marca.upper()
            
            # Filtro 1: LINEA COPIA debe estar vacía o ser INDETERMINADO/#N/D
            if linea and linea_upper not in ("INDETERMINADO", "#N/D", "#N/A", "N/A", "NA", "NONE"):
                continue
            
            # Filtro 2: MARCA SISTEMA debe estar en marcas propias
            if marca_upper not in marcas_propias:
                continue
            
            # Si pasa ambos filtros, agregar a procesar
            filas_a_procesar.append((i, ref, marca))
        
        if not filas_a_procesar:
            log("  ℹ No hay registros para procesar después de aplicar filtros")
            return last_row
        
        log(f"  {len(filas_a_procesar)} registros cumplen los filtros")
        
        # FASE 3: ACTUALIZAR CAMPOS
        log("  Fase 3: Actualizando campos...")
        
        gestor_map = distribucion.get('gestor', {})
        clasif_map = distribucion.get('clasificacion', {})
        
        # Construir diccionario de actualizaciones con FILAS EXCEL CORRECTAS
        updates = {}
        for idx, ref, marca in filas_a_procesar:
            fila_excel = start_data_row + idx  # Esta es la fila DESPUÉS de eliminar
            marca_upper = marca.upper().strip()
            
            updates[fila_excel] = {
                'marca': marca,
                'inv_bodega': "0",
                'linea': marca,
                'sublinea': marca,
                'lider': gestor_map.get(marca_upper, ""),
                'clasificacion': clasif_map.get(marca_upper, "")
            }
        
        # Ordenar filas para actualizaciones eficientes
        filas_ordenadas = sorted(updates.keys())
        
        # Estadísticas
        lideres_encontrados = sum(1 for v in updates.values() if v['lider'])
        clasif_encontradas = sum(1 for v in updates.values() if v['clasificacion'])
        
        log(f"  Actualizando {len(updates)} registros ({lideres_encontrados} con líder, {clasif_encontradas} con clasificación)...")
        
        # Actualizar columnas una por una
        columnas_actualizadas = 0
        
        if col_marca_copia:
            valores = [updates[f]['marca'] for f in filas_ordenadas]
            for i, fila in enumerate(filas_ordenadas):
                ws_inv_copia.Cells(fila, col_marca_copia).Value = valores[i]
            columnas_actualizadas += 1
            log(f"MARCA copia actualizada")
        
        if col_inv_bodega_ger:
            for fila in filas_ordenadas:
                ws_inv_copia.Cells(fila, col_inv_bodega_ger).Value = "0"
            columnas_actualizadas += 1
            log(f"INV BODEGA GERENCIA actualizada")
        
        if col_linea_copia:
            valores = [updates[f]['linea'] for f in filas_ordenadas]
            for i, fila in enumerate(filas_ordenadas):
                ws_inv_copia.Cells(fila, col_linea_copia).Value = valores[i]
            columnas_actualizadas += 1
            log(f"LINEA COPIA actualizada")
        
        if col_sublinea_copia:
            valores = [updates[f]['sublinea'] for f in filas_ordenadas]
            for i, fila in enumerate(filas_ordenadas):
                ws_inv_copia.Cells(fila, col_sublinea_copia).Value = valores[i]
            columnas_actualizadas += 1
            log(f"SUB-LINEA COPIA actualizada")
        
        if col_lider_linea:
            valores = [updates[f]['lider'] for f in filas_ordenadas]
            for i, fila in enumerate(filas_ordenadas):
                ws_inv_copia.Cells(fila, col_lider_linea).Value = valores[i]
            columnas_actualizadas += 1
            log(f"LIDER LINEA actualizada ({lideres_encontrados} valores)")
        
        if col_clasificacion:
            valores = [updates[f]['clasificacion'] for f in filas_ordenadas]
            for i, fila in enumerate(filas_ordenadas):
                ws_inv_copia.Cells(fila, col_clasificacion).Value = valores[i]
            columnas_actualizadas += 1
            log(f"CLASIFICACION actualizada ({clasif_encontradas} valores)")
        
        log(f" Proceso completado: {columnas_actualizadas} columnas actualizadas en {len(updates)} registros")
        
        return last_row
        
    except Exception as e:
        log(f"  ⚠ ERROR al aplicar reglas de marcas propias: {e}")
        import traceback
        log(traceback.format_exc())
        return last_row
    

def eliminar_registros_linea_copia_indeterminada(wsinvcopia, startdatarow: int, lastrow: int, 
                                                  refcolidx: int, hdrncopia: dict) -> int:
    """
    Elimina los registros donde LINEA COPIA tenga valores indeterminados (#N/D).
    """
    try:
        
        # Buscar columna LINEA COPIA usando las claves normalizadas del diccionario
        collineacopia = hdrncopia.get(_norm("LINEA COPIA"))
        
        if not collineacopia:
            log("  ⚠ Columna LINEA COPIA no encontrada")
            return lastrow
        
        # Leer datos de REFERENCIA y LINEA COPIA
        colstoread = [refcolidx, collineacopia]
        data = read_multiple_columns_optimized(wsinvcopia, startdatarow, lastrow, colstoread)
        
        referencias = data[refcolidx]
        lineascopia = data[collineacopia]
        
        # Identificar filas a eliminar
        filas_a_eliminar = []
        
        log(f"Analizando {len(referencias)} registros...")
        
        for i in range(len(referencias)):
            ref = str(referencias[i]).strip() if referencias[i] not in [None, "", "None"] else ""
            linea = str(lineascopia[i]).strip() if lineascopia[i] not in [None, "", "None"] else ""
            linea_upper = linea.upper()
            
            # Filtrar los indeterminados: #N/D, N/A, NA, etc.
            if linea_upper in ["INDETERMINADO", "#N/D", "N/D", "NA", "N/A", "#N/A", "NONE", ""]:
                filas_a_eliminar.append((i, ref, linea))
        
        # Eliminar filas
        if filas_a_eliminar:
            
            # Eliminar en orden inverso para no afectar índices
            for idx, ref, linea in sorted(filas_a_eliminar, reverse=True):
                fila_excel = startdatarow + idx
                try:
                    wsinvcopia.Rows(fila_excel).Delete()
                except Exception as e:
                    log(f"    ⚠ Error al eliminar fila {fila_excel} (Ref: {ref}): {e}")
            
            # Actualizar lastrow
            lastrow = lastrow - len(filas_a_eliminar)
            log(f"{len(filas_a_eliminar)} filas eliminadas. Nuevo rango hasta fila {lastrow}")
        else:
            log("No se encontraron registros con LINEA COPIA indeterminada para eliminar")
        
        return lastrow
        
    except Exception as e:
        log(f"  ❌ ERROR al eliminar registros con LINEA COPIA indeterminada: {e}")
        import traceback
        log(traceback.format_exc())
        return lastrow
    
def procesar_existencias_negativas_y_cero(ws_inv_copia, start_data_row: int, last_row: int, 
                                          ref_col_idx: int, hdrn_copia: dict, base_path: Path) -> int:
    """
    Filtra EXISTENCIA (fecha actual) negativos y ceros.
    """
    try:
        log("Procesando existencias negativas...")
        
        # Buscar columnas necesarias
        col_existencia = None
        for name, col in hdrn_copia.items():
            if name.startswith(_norm("EXISTENCIA")):
                col_existencia = col
                break
        
        col_costo_promedio = hdrn_copia.get(_norm("COSTO PROMEDIO"))
        
        if not col_existencia:
            log("  ⚠ Columna EXISTENCIA no encontrada")
            return last_row
        
        if not col_costo_promedio:
            log("  ⚠ Columna COSTO PROMEDIO no encontrada")
        
        # Leer todas las columnas relevantes para el reporte
        cols_reporte = [ref_col_idx, col_existencia]
        col_nombre_myr = hdrn_copia.get(_norm("NOMBRE MYR"))
        col_marca_copia = hdrn_copia.get(_norm("MARCA COPIA"))
        col_linea_copia = hdrn_copia.get(_norm("LINEA COPIA"))
        
        if col_nombre_myr:
            cols_reporte.append(col_nombre_myr)
        if col_marca_copia:
            cols_reporte.append(col_marca_copia)
        if col_linea_copia:
            cols_reporte.append(col_linea_copia)
        if col_costo_promedio:
            cols_reporte.append(col_costo_promedio)
        
        data = read_multiple_columns_optimized(ws_inv_copia, start_data_row, last_row, cols_reporte)
        
        referencias = data[ref_col_idx]
        existencias = data[col_existencia]
        
        # Identificar SOLO registros negativos
        registros_negativos = []
        
        log(f" Analizando {len(referencias)} registros...")
        
        for i in range(len(referencias)):
            ref = str(referencias[i]).strip() if referencias[i] not in [None, "", "None"] else ""
            
            try:
                exist_val = float(existencias[i]) if existencias[i] not in [None, "", "None"] else 0.0
            except:
                exist_val = 0.0
            
            # SOLO identificar negativos (ignorar ceros)
            if exist_val < 0:
                # Limpiar la referencia para eliminar .0 innecesario
                ref_limpia = ref
                try:
                    # Si es un número entero con .0, quitarlo
                    if '.' in ref and ref.replace('.', '').replace('-', '').isdigit():
                        num_float = float(ref)
                        if abs(num_float - int(num_float)) < 1e-9:  # Es entero
                            ref_limpia = str(int(num_float))
                except:
                    pass  # Mantener ref original si no se puede convertir
                
                registro = {
                    'indice': i,
                    'referencia': ref_limpia,  # ← Usar la referencia limpia
                    'existencia': exist_val
                }
                
                if col_nombre_myr:
                    registro['nombre'] = data[col_nombre_myr][i]
                if col_marca_copia:
                    registro['marca'] = data[col_marca_copia][i]
                if col_linea_copia:
                    registro['linea'] = data[col_linea_copia][i]
                if col_costo_promedio:
                    registro['costo'] = data[col_costo_promedio][i]
                
                registros_negativos.append(registro)                
                   
        # Generar Excel con registros negativos
        if registros_negativos:
            log(f" Se encontraron {len(registros_negativos)} registros con EXISTENCIA NEGATIVA")
            
            try:
                # Crear DataFrame para exportar
                df_negativos = pd.DataFrame(registros_negativos)
                
                # Renombrar columnas
                rename_map = {
                    'referencia': 'REFERENCIA',
                    'existencia': 'EXISTENCIA',
                    'nombre': 'NOMBRE',
                    'marca': 'MARCA',
                    'linea': 'LINEA',
                    'costo': 'COSTO PROMEDIO'
                }
                df_negativos = df_negativos.rename(columns=rename_map)
                
                # Eliminar columna de índice
                if 'indice' in df_negativos.columns:
                    df_negativos = df_negativos.drop(columns=['indice'])
                
                # Generar nombre del archivo
                fecha_actual = datetime.now().strftime("%Y%m%d_%H%M")
                nombre_reporte = f"REPORTE_EXISTENCIAS_NEGATIVAS_{fecha_actual}.xlsx"
                ruta_reporte = base_path / nombre_reporte
                
                # Guardar Excel
                df_negativos.to_excel(ruta_reporte, index=False, engine='openpyxl')
                log(f"Reporte generado: {nombre_reporte}")
                log(f" 📁 Ubicación: {ruta_reporte}")
                
                # Mostrar ejemplos
                for reg in registros_negativos:
                    log(f"    - Ref: {reg['referencia']}, EXISTENCIA: {reg['existencia']}")
                    
            except Exception as e:
                log(f"  ⚠ Error al generar reporte de negativos: {e}")
                import traceback
                log(traceback.format_exc())
            
            # Cambiar a 0 SOLO los registros negativos
            log(f"Cambiando a 0: {len(registros_negativos)} registros negativos")
            
            # Modificar EXISTENCIA y COSTO PROMEDIO solo para negativos
            for reg in registros_negativos:
                fila_excel = start_data_row + reg['indice']
                try:
                    # Cambiar EXISTENCIA a 0
                    ws_inv_copia.Cells(fila_excel, col_existencia).Value = 0
                    
                    # Cambiar COSTO PROMEDIO a 0 si existe
                    if col_costo_promedio:
                        ws_inv_copia.Cells(fila_excel, col_costo_promedio).Value = 0
                        
                except Exception as e:
                    log(f"    ⚠ Error al actualizar fila {fila_excel} (Ref: {reg['referencia']}): {e}")
            
            log(f"{len(registros_negativos)} registros negativos actualizados a 0")
        else:
            log("No se encontraron registros con EXISTENCIA negativa")
        
        return last_row
        
    except Exception as e:
        log(f"  ❌ ERROR al procesar existencias negativas: {e}")
        import traceback
        log(traceback.format_exc())
        return last_row


# ==== EXCEL COM ====
def excel_open(path: Path, password: str | None = None):
    """Abre con COM en modo silencioso y OPTIMIZADO."""
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.Interactive = False
    excel.EnableEvents = False
    excel.ScreenUpdating = False
    
    try: excel.AskToUpdateLinks = False
    except Exception: pass
    try: excel.AutomationSecurity = 3
    except Exception: pass

    info = {"tmp_path": None, "target_path": str(path), "reapply_password": None}

    encrypted = False
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            with open(path, "rb") as f:
                of = msoffcrypto.OfficeFile(f)
                encrypted = bool(getattr(of, "is_encrypted", False))
        except Exception:
            encrypted = False

    src_path = path
    if encrypted:
        ok = False
        for pw in PASSWORDS_TRY:
            try:
                bio = decrypt_to_stream(path, pw)
                tmp = save_bytesio_to_temp(bio, Path(path).stem)
                src_path = tmp
                info["tmp_path"] = str(tmp)
                info["reapply_password"] = pw
                ok = True
                break
            except Exception:
                continue
        if not ok:
            excel.Quit()
            raise RuntimeError(f"El libro '{path.name}' está cifrado y ninguna contraseña funcionó.")

    try:
        wb = excel.Workbooks.Open(str(src_path), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
        try:
            excel.Calculation = -4135  
        except Exception as e:
            log(f"Aviso: no se pudo establecer cálculo manual: {e}")
        return excel, wb, info
    except Exception as e:
        excel.Quit()
        raise RuntimeError(f"No pude abrir el libro {path.name} de forma silenciosa.") from e

def excel_close(excel, wb, save=True):
    try:
        if save:
            excel.Calculation = -4105  
        wb.Close(SaveChanges=save)
    finally:
        excel.Quit()

def ws_headers(ws, header_row_visible: int) -> tuple[dict, dict]:
    """Devuelve (mapa header→col_idx, mapa normalizado→col_idx)"""
    used_cols = ws.UsedRange.Columns.Count
    hdr = {}
    for c in range(1, used_cols+1):
        v = ws.Cells(header_row_visible, c).Value
        if v is None: continue
        s = str(v).strip()
        if s and s != "None":
            hdr[s] = c
    hdrn = {_norm(k): v for k, v in hdr.items()}
    return hdr, hdrn

# ==== AJUSTES PIVOT-SAFE ====
def ws_first_pivot_row(ws) -> int | None:
    """Fila superior de la primera PivotTable, o None si no hay."""
    try:
        pts = ws.PivotTables()
        count = int(getattr(pts, "Count", 0))
        if count == 0:
            return None
        first = None
        for i in range(1, count + 1):
            try:
                r = pts(i).TableRange2.Row
                if first is None or r < first:
                    first = r
            except Exception:
                pass
        return first
    except Exception:
        return None

def ws_pivot_blocks(ws):
    """Lista de bloques de pivots [(r1, r2, c1, c2), ...]."""
    blocks = []
    try:
        pts = ws.PivotTables()
        count = int(getattr(pts, "Count", 0))
        for i in range(1, count + 1):
            try:
                tr = pts(i).TableRange2
                r1, c1 = tr.Row, tr.Column
                r2 = r1 + tr.Rows.Count - 1
                c2 = c1 + tr.Columns.Count - 1
                blocks.append((int(r1), int(r2), int(c1), int(c2)))
            except Exception:
                pass
    except Exception:
        pass
    return blocks

def ws_ensure_range(ws, start_row: int, expected_rows: int, header_row: int) -> int:
    """
    Asegura que el rango detectado incluya todas las filas esperadas.
    """
    calculated_last = start_row + expected_rows - 1
    
    pivot_top = ws_first_pivot_row(ws)
    if pivot_top and pivot_top > header_row:
        if calculated_last >= pivot_top:
            log(f"⚠ Límite por pivot: reduciendo de {calculated_last} a {pivot_top - 1}")
            return pivot_top - 1
    
    return calculated_last


def ws_apply_borders_to_range(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    """
    Aplica bordes a un rango completo de celdas.
    """
    try:
        
        full_range = ws.Range(ws.Cells(start_row, start_col), ws.Cells(end_row, end_col))
        
        for border_id in [7, 8, 9, 10, 11, 12]:
            try:
                full_range.Borders(border_id).LineStyle = 1      
                full_range.Borders(border_id).Weight = 2         
                full_range.Borders(border_id).ColorIndex = -4105 
            except Exception:
                continue
            
    except Exception as e:
        log(f"  ⚠ Error al aplicar bordes: {e}")


def ws_remove_formatting_from_range(ws, start_row: int, end_row: int, start_col: int, end_col: int):
    """
    Elimina formato de negrita y color de fondo de un rango.
    Preserva los formatos de número (General, Contabilidad, etc.) de cada columna.
    """
    try:
        
        for col in range(start_col, end_col + 1):
            try:
                original_number_format = ws.Cells(start_row, col).NumberFormat

                col_range = ws.Range(ws.Cells(start_row, col), ws.Cells(end_row, col))
                
                try:
                    col_range.Font.Bold = False
                except Exception:
                    pass
                
                try:
                    col_range.Interior.ColorIndex = 0  
                except Exception:
                    pass
                
                try:
                    col_range.NumberFormat = original_number_format
                except Exception:
                    pass
                    
            except Exception:
                continue
        
       
    except Exception as e:
        log(f"  ⚠ Error al limpiar formato: {e}")


def ws_update_subtotal_formula(ws, formula_row: int, last_data_row: int):
    """Actualiza la fórmula de subtotal en la fila 1 para que abarque todo el rango."""
    try:
        
        used_cols = ws.UsedRange.Columns.Count
        updated_count = 0
        
        for col in range(1, used_cols + 1):
            try:
                cell_formula = ws.Cells(formula_row, col).Formula
                
                if cell_formula and "SUBTOTAL" in str(cell_formula).upper():
                    import re
                    match = re.search(r'SUBTOTAL\((\d+),', str(cell_formula))
                    
                    if match:
                        func_num = match.group(1)
                        col_letter = _col_num_to_letter(col)
                        new_formula = f"=SUBTOTAL({func_num},{col_letter}3:{col_letter}{last_data_row})"
                        ws.Cells(formula_row, col).Formula = new_formula
                        updated_count += 1
                        
            except Exception:
                continue      
        
    except Exception as e:
        log(f"  ⚠ Error al actualizar fórmulas de subtotal: {e}")


def ws_add_final_subtotals(ws, last_data_row: int, header_row: int, hdrn: dict):
    """
    Agrega subtotales al final de todos los registros para EXISTENCIA y TOTAL INV.
    También agrega subtotales en G1 e I1.
    Usa funciones SUBTOTAL compatibles con filtros dinámicos.
    """
    try:
        log(f"Agregando subtotales finales en fila {last_data_row + 1}...")
        
        subtotal_row = last_data_row + 1
        
        # Buscar columna EXISTENCIA
        exist_col = None
        for name, col in hdrn.items():
            if name.startswith("existencia "):
                exist_col = col
                break
        
        total_inv_col = hdrn.get(_norm("TOTAL INV"))      
        header_color = None
        try:
            header_color = ws.Cells(header_row, 1).Interior.Color
        except Exception:
            header_color = 15849925  
        
        subtotals_added = 0

        # Subtotal EXISTENCIA - Usar función 109 para SUMA (ignora filas ocultas)
        if exist_col:
            try:
                col_letter = _col_num_to_letter(exist_col)
                # 109 = SUMA ignorando filas ocultas por filtros
                formula = f"=SUBTOTAL(109,{col_letter}{header_row + 1}:{col_letter}{last_data_row})"
                
                cell = ws.Cells(subtotal_row, exist_col)               
                cell.Formula = formula
                
                # Formato sin decimales y con punto como separador de miles
                cell.NumberFormat = "#.##0"
                cell.Font.Bold = True
                cell.Interior.Color = header_color
                
                try:
                    for border_id in [7, 8, 9, 10]:
                        cell.Borders(border_id).LineStyle = 1
                        cell.Borders(border_id).Weight = 2
                        cell.Borders(border_id).ColorIndex = -4105
                except Exception:
                    pass
                
                subtotals_added += 1
                log(f"Subtotal EXISTENCIA agregado en fila {subtotal_row} (formato: #.##0)")
                
                # Agregar subtotal en G1 SIN FONDO AZUL, solo negrilla
                try:
                    cell_g1 = ws.Cells(1, exist_col)
                    cell_g1.Formula = formula
                    cell_g1.NumberFormat = "#.##0"
                    cell_g1.Font.Bold = True
                    cell_g1.Interior.ColorIndex = -4142  
                    
                    try:
                        for border_id in [7, 8, 9, 10]:
                            cell_g1.Borders(border_id).LineStyle = 1
                            cell_g1.Borders(border_id).Weight = 2
                            cell_g1.Borders(border_id).ColorIndex = -4105
                    except Exception:
                        pass
                    
                    log(f"Subtotal EXISTENCIA también agregado en G1 (sin fondo, solo negrilla)")
                except Exception as e:
                    log(f"    ⚠ Error al agregar subtotal en G1: {e}")
                
            except Exception as e:
                log(f"    ⚠ Error al agregar subtotal EXISTENCIA: {e}")
                import traceback
                log(traceback.format_exc())
        
        # Subtotal TOTAL INV
        if total_inv_col:
            try:
                col_letter = _col_num_to_letter(total_inv_col)
                formula = f"=SUBTOTAL(109,{col_letter}{header_row + 1}:{col_letter}{last_data_row})"
                
                cell = ws.Cells(subtotal_row, total_inv_col)               
                cell.Formula = formula
                cell.Font.Bold = True
                cell.Interior.Color = header_color

                try:
                    # Copiar el formato de la última fila de datos
                    original_format = ws.Cells(last_data_row, total_inv_col).NumberFormat
                    cell.NumberFormat = original_format
                except Exception as e:
                    log(f"    ⚠ No se pudo copiar formato original: {e}")
                    # Formato por defecto si falla (contabilidad con 2 decimales)
                    cell.NumberFormat = "_($* #,##0.00_);_($* (#,##0.00);_($* \"-\"??_);_(@_)"

                try:
                    for border_id in [7, 8, 9, 10]:
                        cell.Borders(border_id).LineStyle = 1
                        cell.Borders(border_id).Weight = 2
                        cell.Borders(border_id).ColorIndex = -4105
                except Exception:
                    pass
                
                subtotals_added += 1
                log(f"Subtotal TOTAL INV agregado en fila {subtotal_row}")
                
                #Agregar subtotal en I1 con fondo AMARILLO
                try:
                    cell_i1 = ws.Cells(1, total_inv_col)
                    cell_i1.Formula = formula
                    cell_i1.Font.Bold = True
                    
                    #Color amarillo (65535 en RGB o 6 en ColorIndex)
                    cell_i1.Interior.Color = 65535  # Amarillo RGB
                    
                    try:
                        original_format = ws.Cells(last_data_row, total_inv_col).NumberFormat
                        cell_i1.NumberFormat = original_format
                    except Exception:
                        cell_i1.NumberFormat = "_($* #,##0.00_);_($* (#,##0.00);_($* \"-\"??_);_(@_)"
                    
                    try:
                        for border_id in [7, 8, 9, 10]:
                            cell_i1.Borders(border_id).LineStyle = 1
                            cell_i1.Borders(border_id).Weight = 2
                            cell_i1.Borders(border_id).ColorIndex = -4105
                    except Exception:
                        pass
                    
                    log(f"Subtotal TOTAL INV también agregado en I1 (fondo amarillo)")
                except Exception as e:
                    log(f"    ⚠ Error al agregar subtotal en I1: {e}")
                
            except Exception as e:
                log(f"    ⚠ Error al agregar subtotal TOTAL INV: {e}")
        
        if subtotals_added > 0:
            log(f"{subtotals_added} subtotales agregados con fórmulas dinámicas (compatibles con filtros)")
            log(f"G1: sin fondo (solo negrilla) | I1: fondo amarillo")
        else:
            log(f"  ⚠ No se pudieron agregar subtotales")
        
    except Exception as e:
        log(f"  ⚠ Error al agregar subtotales finales: {e}")
        import traceback
        log(traceback.format_exc())

def _col_num_to_letter(col_num):
    """Convierte número de columna a letra."""
    letter = ''
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter


def _ranges_without_pivots_for_column(col_idx: int, start_row: int, end_row: int, pivot_blocks):
    """Devuelve sub-rangos [a,b] dentro de [start_row,end_row] que NO cruzan pivots."""
    if end_row < start_row:
        return []
    holes = []
    for (r1, r2, c1, c2) in pivot_blocks:
        if c1 <= col_idx <= c2:
            holes.append((max(r1, start_row), min(r2, end_row)))
    holes.sort()
    segments = []
    cur = start_row
    for (h1, h2) in holes:
        if h2 < cur or h1 > end_row:
            continue
        if h1 > cur:
            segments.append((cur, h1 - 1))
        cur = max(cur, h2 + 1)
    if cur <= end_row:
        segments.append((cur, end_row))
    return [(a, b) for (a, b) in segments if b >= a]


def aplicar_autofiltros_y_ordenar(ws, header_row: int, last_row: int, hdrn: dict):
    """
    Aplica autofiltros a todos los encabezados y ordena por TOTAL INV de mayor a menor.
    """
    try:
        log("APLICANDO AUTOFILTROS Y ORDENAMIENTO...")
        
        # Buscar columna TOTAL INV
        total_inv_col = hdrn.get(_norm("TOTAL INV"))
        
        if not total_inv_col:
            log("  ⚠ Columna TOTAL INV no encontrada")
            log(f"  Columnas disponibles: {list(hdrn.keys())}")
            return
        
        log(f"  ✓ Columna TOTAL INV encontrada: índice {total_inv_col}")
        
        # Determinar el rango completo para el autofiltro
        used_range = ws.UsedRange
        first_col = used_range.Column
        last_col = first_col + used_range.Columns.Count - 1
        
        log(f"  → Rango de datos: filas {header_row} a {last_row}, columnas {first_col} a {last_col}")
        
        # Crear el rango del encabezado
        header_range = ws.Range(
            ws.Cells(header_row, first_col),
            ws.Cells(last_row, last_col)
        )
        
        # Eliminar AutoFilter existente
        try:
            if ws.AutoFilterMode:
                ws.AutoFilterMode = False
                log("  ✓ AutoFilter anterior eliminado")
        except Exception as e:
            log(f"  ⚠ Error al eliminar AutoFilter: {e}")
        

        # Ordenar por TOTAL INV de MAYOR A MENOR
        try:
            log(f"  → Preparando ordenamiento por columna {total_inv_col} (TOTAL INV)...")
            
            # Crear la clave de ordenamiento
            sort_key = ws.Cells(header_row, total_inv_col)
            
            # Aplicar el ordenamiento usando Sort
            header_range.Sort(
                Key1=sort_key,
                Order1=2,         # xlDescending (2 = descendente)
                Header=1,         # xlYes (1 = tiene encabezado)
                MatchCase=False,
                Orientation=1     # xlTopToBottom
            )
            
            log("  ✓ Datos ordenados por TOTAL INV (MAYOR A MENOR)")
            
            # Verificar el ordenamiento leyendo las primeras 5 filas
            log("  → Verificando orden (primeras 5 filas):")
            for row in range(header_row + 1, min(header_row + 6, last_row + 1)):
                try:
                    valor = ws.Cells(row, total_inv_col).Value
                    if valor is not None:
                        log(f"     Fila {row}: {valor:,.2f}" if isinstance(valor, (int, float)) else f"     Fila {row}: {valor}")
                except Exception:
                    pass
                    
        except Exception as e:
            log(f"  ❌ ERROR al ordenar por TOTAL INV: {e}")
            import traceback
            log(traceback.format_exc())
            return
        
        log("✓ Autofiltros y ordenamiento completados exitosamente")
        log("")
        
    except Exception as e:
        log(f"❌ ERROR CRÍTICO al aplicar autofiltros y ordenar: {e}")
        import traceback
        log(traceback.format_exc())


def eliminar_registros_estandarizados(ws_inv_copia, start_data_row: int, last_row: int, 
                                      ref_col_idx: int, hdrn_copia: dict) -> int:
    """
    Elimina registros según criterios estandarizados:
    1. NOMBRE MYR que contenga "Publicidad"
    2. REFERENCIA con patrón: 2 ceros + 2 números + letras (SIN símbolos como /)
    3. NOMBRE MYR que contenga "Ajuste de precios"
    
    Retorna el nuevo last_row después de las eliminaciones.
    """
    try:
        log("="*60)
        log("ELIMINANDO REGISTROS SEGÚN CRITERIOS ESTANDARIZADOS")
        log("="*60)
        
        # Buscar columnas necesarias
        col_nombre_myr = hdrn_copia.get(_norm("NOMBRE MYR"))
        
        if not col_nombre_myr:
            log("  ⚠ Columna NOMBRE MYR no encontrada")
            return last_row
        
        # Leer datos de REFERENCIA y NOMBRE MYR
        cols_to_read = [ref_col_idx, col_nombre_myr]
        data = read_multiple_columns_optimized(ws_inv_copia, start_data_row, last_row, cols_to_read)
        
        referencias = data[ref_col_idx]
        nombres_myr = data[col_nombre_myr]
        
        # Identificar filas a eliminar según los 3 criterios
        filas_a_eliminar = []
        
        log(f"Analizando {len(referencias)} registros...")
        
        for i in range(len(referencias)):
            ref = str(referencias[i]).strip() if referencias[i] not in [None, "", "None"] else ""
            nombre = str(nombres_myr[i]).strip() if nombres_myr[i] not in [None, "", "None"] else ""
            
            motivo_eliminacion = None
            
            # CRITERIO 1: NOMBRE MYR contiene "Publicidad"
            if "publicidad" in nombre.lower():
                motivo_eliminacion = "Publicidad en NOMBRE MYR"
            
            # CRITERIO 2: REFERENCIA con patrón 00##Letras (sin símbolos como /)
            # Ejemplo: 0041R, 0012AB, etc.
            elif ref and "/" not in ref and "\\" not in ref:
                # Verificar patrón: 2 ceros iniciales + 2 dígitos + letras
                import re
                # Patrón: exactamente 2 ceros, seguido de exactamente 2 dígitos, seguido de al menos una letra
                patron = r'^00\d{2}[A-Za-z]+$'
                if re.match(patron, ref):
                    motivo_eliminacion = "Patrón 00##Letras en REFERENCIA"
            
            # CRITERIO 3: NOMBRE MYR contiene "Ajuste de precios"
            if not motivo_eliminacion and "ajuste de precios" in nombre.lower():
                motivo_eliminacion = "Ajuste de precios en NOMBRE MYR"
            
            # Si cumple algún criterio, agregar a lista de eliminación
            if motivo_eliminacion:
                filas_a_eliminar.append((i, ref, nombre, motivo_eliminacion))
        
        # Eliminar filas
        if filas_a_eliminar:
            log(f"")
            log(f"Se encontraron {len(filas_a_eliminar)} registros para eliminar:")
            log(f"")
            
            # Mostrar resumen por motivo
            from collections import Counter
            motivos = Counter([item[3] for item in filas_a_eliminar])
            for motivo, cantidad in motivos.items():
                log(f"  • {motivo}: {cantidad} registros")
            
            log(f"")
            log(f"Ejemplos de registros a eliminar:")
            for idx, ref, nombre, motivo in filas_a_eliminar[:10]:
                log(f"  - Ref: {ref[:30]:<30} | Nombre: {nombre[:40]:<40} | Motivo: {motivo}")
            
            if len(filas_a_eliminar) > 10:
                log(f"  ... y {len(filas_a_eliminar) - 10} más")
            
            log(f"")
            log(f"Eliminando filas...")
            
            # Eliminar en orden inverso para no afectar índices
            for idx, ref, nombre, motivo in sorted(filas_a_eliminar, reverse=True):
                fila_excel = start_data_row + idx
                try:
                    ws_inv_copia.Rows(fila_excel).Delete()
                except Exception as e:
                    log(f"    ⚠ Error al eliminar fila {fila_excel} (Ref: {ref}): {e}")
            
            # Actualizar last_row
            last_row = last_row - len(filas_a_eliminar)
            log(f"")
            log(f"✓ {len(filas_a_eliminar)} filas eliminadas exitosamente")
            log(f"✓ Nuevo rango: hasta fila {last_row}")
            log(f"")
        else:
            log("✓ No se encontraron registros que cumplan los criterios de eliminación")
            log("")
        
        return last_row
        
    except Exception as e:
        log(f"❌ ERROR al eliminar registros estandarizados: {e}")
        import traceback
        log(traceback.format_exc())
        return last_row


# ==== WS UTILS  ====
def ws_last_row(ws, key_col_idx: int, header_row_visible: int):
    """Última fila con datos."""
    last = ws.Cells(ws.Rows.Count, key_col_idx).End(-4162).Row
    return max(last, header_row_visible)

def ws_fill_column_values(ws, col_idx: int, start_row: int, values: list):
    """Escribe valores en una columna saltando pivots."""
    n = len(values)
    if n == 0:
        return

    end_row = start_row + n - 1
    pivots = ws_pivot_blocks(ws)
    safe_segments = _ranges_without_pivots_for_column(col_idx, start_row, end_row, pivots)

    offset = 0
    for (a, b) in safe_segments:
        if offset >= n:
            break
        seg_len = min(b - a + 1, n - offset)
        if seg_len <= 0:
            continue

        chunk = values[offset: offset + seg_len]
        chunk = [("" if (v is None or (isinstance(v, float) and np.isnan(v))) else v) for v in chunk]

        rng = ws.Range(ws.Cells(a, col_idx), ws.Cells(a + seg_len - 1, col_idx))
        rng.Value = [[v] for v in chunk]
        offset += seg_len

def ws_clear_column(ws, col_idx: int, start_row: int, end_row: int):
    """Limpia una columna por tramos, evitando pivots."""
    if end_row < start_row:
        return
    pivots = ws_pivot_blocks(ws)
    safe_segments = _ranges_without_pivots_for_column(col_idx, start_row, end_row, pivots)
    for (a, b) in safe_segments:
        rng = ws.Range(ws.Cells(a, col_idx), ws.Cells(b, col_idx))
        rng.ClearContents()

def ws_copy_down_formula(ws, col_idx: int, start_row: int, end_row: int):
    """Copia la fórmula desde start_row hasta end_row."""
    if end_row < start_row: return
    fml = ws.Cells(start_row, col_idx).Formula
    if fml:
        rng = ws.Range(ws.Cells(start_row, col_idx), ws.Cells(end_row, col_idx))
        rng.Formula = fml

def ws_headers_smart(ws, expected_row: int, preferred_labels: list[str] | None = None):
    """Detecta de forma robusta la fila de encabezado."""
    preferred_norm = [_norm(x) for x in (preferred_labels or [])]
    tried = [expected_row] + [r for r in range(1, 11) if r != expected_row]
    for hr in tried:
        hdr, hdrn = ws_headers(ws, hr)
        if not hdrn:
            continue
        if not preferred_norm or any(lbl in hdrn for lbl in preferred_norm):
            return hr, hdr, hdrn
    try:
        first_row = ws.UsedRange.Row
        hdr, hdrn = ws_headers(ws, first_row)
        if hdrn:
            return first_row, hdr, hdrn
    except Exception:
        pass
    return expected_row, {}, {}

def find_reference_col_idx(hdrn: dict, ws, header_row_used: int) -> int:
    """Encuentra índice de columna para REFERENCIA."""
    for name in ["REFERENCIA", "REFERENCIA FERTRAC", "REFERENCIA INTERNA", "REF", "CÓDIGO", "CODIGO", "SKU"]:
        cidx = hdrn.get(_norm(name))
        if cidx:
            return cidx
    for k, v in hdrn.items():
        if "referenc" in k or "codigo" in k or k.endswith("ref"):
            return v
    used_cols = ws.UsedRange.Columns.Count
    for c in range(1, used_cols + 1):
        for r in range(header_row_used + 1, header_row_used + 15):
            val = ws.Cells(r, c).Value
            if val not in (None, "", "None"):
                return c
    return 1

def ws_ensure_existencia_header(ws, header_row_visible: int) -> int:
    """Devuelve col_idx del encabezado EXISTENCIA {MES DD}."""
    target = exist_col_title_for_today()
    hdr, hdrn = ws_headers(ws, header_row_visible)
    target_col = None
    for name, col in hdr.items():
        if _norm(name).startswith("existencia "):
            target_col = col
            ws.Cells(header_row_visible, target_col).Value = target
            break
    if target_col is None:
        used_cols = ws.UsedRange.Columns.Count
        target_col = used_cols + 1
        ws.Cells(header_row_visible, target_col).Value = target
    return target_col

def normalize_sheet_name(wb, target_name: str) -> str:
    """Normaliza el nombre de una hoja eliminando espacios extras."""
    target_norm = _norm(target_name)
    
    for i in range(1, wb.Worksheets.Count + 1):
        ws = wb.Worksheets(i)
        sheet_name = ws.Name
        sheet_norm = _norm(sheet_name)
        
        if sheet_norm == target_norm or target_norm in sheet_norm:
            clean_name = sheet_name.strip()
            if clean_name != sheet_name:
                try:
                    ws.Name = clean_name
                    log(f"Nombre de hoja normalizado: '{sheet_name}' → '{clean_name}'")
                    return clean_name
                except Exception as e:
                    log(f"No se pudo renombrar hoja: {e}")
                    return sheet_name
            return clean_name
    
    return target_name

def read_range_as_array(ws, start_row: int, end_row: int, col_idx: int):
    """Lee un rango completo en una sola operación."""
    if end_row < start_row:
        return []
    rng = ws.Range(ws.Cells(start_row, col_idx), ws.Cells(end_row, col_idx))
    values = rng.Value
    if values is None:
        return [""] * (end_row - start_row + 1)
    if not isinstance(values, (list, tuple)):
        return [values]
    return [row[0] if isinstance(row, (list, tuple)) else row for row in values]

def write_range_as_array(ws, start_row: int, col_idx: int, values: list):
    """Escribe un rango completo en una sola operación."""
    if not values:
        return
    end_row = start_row + len(values) - 1
    rng = ws.Range(ws.Cells(start_row, col_idx), ws.Cells(end_row, col_idx))
    rng.Value = [[v] for v in values]

def read_multiple_columns_optimized(ws, start_row: int, end_row: int, col_indices: list[int]) -> dict:
    """Lee múltiples columnas en UNA SOLA operación - OPTIMIZACIÓN CRÍTICA."""
    if end_row < start_row or not col_indices:
        return {col: [] for col in col_indices}
    
    min_col = min(col_indices)
    max_col = max(col_indices)
    
    rng = ws.Range(ws.Cells(start_row, min_col), ws.Cells(end_row, max_col))
    values = rng.Value
    
    if values is None:
        return {col: [""] * (end_row - start_row + 1) for col in col_indices}
    
    if not isinstance(values[0], (list, tuple)):
        values = [values]
    
    result = {}
    for col_idx in col_indices:
        offset = col_idx - min_col
        result[col_idx] = [row[offset] if isinstance(row, (list, tuple)) and len(row) > offset else "" 
                          for row in values]
    
    return result

# ==== PROCESO PRINCIPAL ====
def main():
    if not HAS_COM:
        raise RuntimeError("Este script requiere Excel COM (win32com). Instálalo y ejecuta en Windows con Excel.")

    log("== Inicio actualización de inventario ==")

    # 1) Cargar datos externos
    log("Cargando datos externos...")

    df_src = cargar_inventario_actualizado(BASE_PATH)

    # Valorizados
    df_val_gen   = cargar_valorizado(BASE_PATH, PFX_VAL_GENERAL)
    df_val_impo  = cargar_valorizado(BASE_PATH, PFX_VAL_FALT_IMPO)

    # FIX: Buscar VALORIZADO FALTANTES sin IMPO
    log("Buscando VALORIZADO FALTANTES (excluyendo IMPO)...")
    archivo_faltantes = None
    for f in BASE_PATH.iterdir():
        if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm'):
            nombre_sin_simbolos = _strip_dol_tmp(f.name)
            nombre_normalizado = _norm(nombre_sin_simbolos)
            
            # Debe ser exactamente "VALORIZADO FALTANTES" (sin IMPO)
            if nombre_normalizado == _norm("VALORIZADO FALTANTES"):
                archivo_faltantes = f
                log(f"  → Encontrado: {f.name}")
                break

    if archivo_faltantes:
        df_val_falt = cargar_valorizado_desde_ruta(archivo_faltantes)
    else:
        log(f"  ⚠️ No encontrado, usando vacío")
        df_val_falt = pd.DataFrame(columns=["__REF_INT__", "__CANT__"])

    df_val_tob   = cargar_valorizado(BASE_PATH, PFX_VAL_TOBERIN)
    
    # Cargar Matriz USD
    df_matriz_usd = cargar_matriz_usd(BASE_PATH)
    matriz_map = df_matriz_usd.set_index("__REF_MATRIZ__")["__DESC_LISTA__"].to_dict() if len(df_matriz_usd) > 0 else {}
    log(f"Matriz USD: {len(matriz_map)} referencias disponibles para actualizar NOMBRE LISTA")

    #Crear diccionario para REFERENCIA LISTA DE PRECIOS
    matriz_map_ref_lista = df_matriz_usd.set_index("__REF_MATRIZ__")["__REF_LISTA_PRECIOS__"].to_dict() if len(df_matriz_usd) > 0 else {}
    if len(matriz_map_ref_lista) > 0:
        no_vacias = sum(1 for v in matriz_map_ref_lista.values() if v and str(v).strip() not in ("", "0", "None"))
        log(f"Matriz USD: {no_vacias} referencias de lista de precios disponibles")

    # Cargar archivos auxiliares para marcas propias
    marcas_propias = cargar_marcas(BASE_PATH)
    log(f"Marcas propias: {len(marcas_propias)} marcas cargadas")
    
    distribucion = cargar_distribucion(BASE_PATH)
    log(f"Distribución: {len(distribucion['gestor'])} gestores, {len(distribucion['clasificacion'])} clasificaciones")

    # Cargar Mayor Existencia
    df_mayor_exist = cargar_mayor_existencia(BASE_PATH)
    mayor_exist_map = df_mayor_exist.set_index("__REF_MAYOR__")["__REM_CONSIG__"].to_dict() if len(df_mayor_exist) > 0 else {}
    if len(mayor_exist_map) > 0:
        no_cero = sum(1 for v in mayor_exist_map.values() if v != 0)
        log(f"Mayor Existencia: {no_cero} referencias con REM EN CONSIG diferente de cero")

    # Join de cantidades
    val_map_impo = df_val_impo.set_index("__REF_INT__")["__CANT__"]
    val_map_falt = df_val_falt.set_index("__REF_INT__")["__CANT__"]
    val_map_tob  = df_val_tob.set_index("__REF_INT__")["__CANT__"]

    # Calcular columnas en VALORIZADO GENERAL
    df_val_gen = df_val_gen.copy()
    df_val_gen["__IMPO_MATCH__"] = df_val_gen["__REF_INT__"].isin(val_map_impo.index)
    df_val_gen["__IMPO_CANT__"]  = df_val_gen["__REF_INT__"].map(val_map_impo).fillna(0.0)
    df_val_gen["__IMPO_DIF__"]   = df_val_gen["__CANT__"] - df_val_gen["__IMPO_CANT__"]

    df_val_gen["__FALT_MATCH__"] = df_val_gen["__REF_INT__"].isin(val_map_falt.index)
    df_val_gen["__FALT_CANT__"]  = df_val_gen["__REF_INT__"].map(val_map_falt).fillna(0.0)
    df_val_gen["__FALT_DIF__"]   = df_val_gen["__CANT__"] - df_val_gen["__FALT_CANT__"]

    df_val_gen["__TOB_MATCH__"]  = df_val_gen["__REF_INT__"].isin(val_map_tob.index)
    df_val_gen["__TOB_CANT__"]   = df_val_gen["__REF_INT__"].map(val_map_tob).fillna(0.0)
    df_val_gen["__TOB_DIF__"]    = df_val_gen["__CANT__"] - df_val_gen["__TOB_CANT__"]
    

    # Consolidado EXISTENCIA_CALC
    # FÓRMULA: VALORIZADO GENERAL - FALTANTES IMPO - FALTANTES - TOBERÍN
    df_val_gen["__EXIST_CALC__"] = (
        df_val_gen["__CANT__"] 
        - df_val_gen["__IMPO_CANT__"] 
        - df_val_gen["__FALT_CANT__"] 
        - df_val_gen["__TOB_CANT__"]
    )
    exist_map = df_val_gen.set_index("__REF_INT__")["__EXIST_CALC__"]


     # 2) Abrir libro PLANTILLA
    log(f"Buscando archivo que coincida con: {PATRON_INV_FILE}")
    p_inv = find_by_prefix(BASE_PATH, PATRON_INV_FILE)
    log(f"Archivo encontrado: {p_inv.name}")
    log(f"Abriendo libro Excel: {p_inv}")
    excel, wb, saveinfo = excel_open(p_inv, password=PASS_INV)

    # ===== NUEVO: ACTUALIZAR REFERENCIAS ANTES DE TODO =====
    log("")
    log("╔" + "="*68 + "╗")
    log("║" + " FASE PREVIA: ACTUALIZANDO REFERENCIAS EN INVENTARIO ORIGINAL ".center(68) + "║")
    log("╚" + "="*68 + "╝")
    log("")

    # Normalizar y abrir hoja INVENTARIO original
    normalized_inv_name = normalize_sheet_name(wb, SHEET_INV_ORIG)

    try:
        ws_inv_orig_temp = wb.Worksheets(normalized_inv_name)
    except Exception:
        ws_inv_orig_temp = wb.Worksheets(1)
        normalized_inv_name = ws_inv_orig_temp.Name

    # EJECUTAR ACTUALIZACIÓN DE REFERENCIAS
    try:
        reemplazos = actualizar_referencias_inventario_original(
            wb, 
            ws_inv_orig_temp, 
            BASE_PATH, 
            PASS_INV
        )
        
        if reemplazos > 0:
            log("")
            log("╔" + "="*68 + "╗")
            log("║" + f" ✅ {reemplazos} REFERENCIAS ACTUALIZADAS EN INVENTARIO ORIGINAL ".center(68) + "║")
            log("║" + " → Las referencias del ERP ahora coincidirán en los cruces ".center(68) + "║")
            log("╚" + "="*68 + "╝")
            log("")
        else:
            log("ℹ️  No se actualizaron referencias (archivo no encontrado o sin cambios)")
            log("")
            
    except Exception as e:
        log(f"⚠️  Error en actualización de referencias: {e}")
        log("   → Continuando con el proceso normal...")
        log("")


    # ===== ABRIR MATRIZ USD EN LA MISMA INSTANCIA DE EXCEL =====
    log("")
    log("ABRIENDO MATRIZ USD en la misma instancia de Excel...")

    matriz_wb = None
    matriz_tmp_path = None

    try:
        # Buscar archivo MATRIZ USD
        log(f"Buscando archivo: {PFX_MATRIZ_USD}")
        matriz_path = find_by_prefix(BASE_PATH, PFX_MATRIZ_USD)
        log(f"  → Encontrado: {matriz_path.name}")
        
        # Verificar si está encriptado
        encrypted = is_encrypted_xlsx(matriz_path)
        
        if encrypted:
            log("  → Archivo encriptado - desencriptando...")
            # Desencriptar a archivo temporal
            ok = False
            for pw in PASSWORDS_TRY:
                try:
                    bio = decrypt_to_stream(matriz_path, pw)
                    tmp = save_bytesio_to_temp(bio, Path(matriz_path).stem)
                    matriz_tmp_path = str(tmp)
                    
                    # CAMBIO CRÍTICO: Abrir con UpdateLinks=3 para desactivar vínculos
                    matriz_wb = excel.Workbooks.Open(
                        str(tmp),
                        UpdateLinks=3,  # ← CAMBIO CLAVE: 3 = No actualizar vínculos externos
                        ReadOnly=True,
                        IgnoreReadOnlyRecommended=True,
                        Password=pw,
                        Notify=False
                    )
                    ok = True
                    log("  ✓ MATRIZ USD abierto (desencriptado) sin actualizar vínculos")
                    break
                except Exception as e:
                    continue
            
            if not ok:
                log("  ⚠ No se pudo desencriptar MATRIZ USD con ninguna contraseña")
        else:
            # Abrir directamente si no está encriptado
            matriz_wb = excel.Workbooks.Open(
                str(matriz_path),
                UpdateLinks=3,  # ← CAMBIO CLAVE: 3 = No actualizar vínculos externos
                ReadOnly=True,
                IgnoreReadOnlyRecommended=True,
                Notify=False
            )
            log("  ✓ MATRIZ USD abierto sin actualizar vínculos")
        
        log("")
        
    except Exception as e:
        log("")
        log("  ⚠ ADVERTENCIA: No se pudo abrir MATRIZ USD")
        log(f"  Motivo: {e}")
        log("  → Puede solicitar contraseña manualmente")
        log("")


    # 3) NORMALIZAR nombre de hoja INVENTARIO
    log("Normalizando nombre de hoja INVENTARIO...")
    normalized_inv_name = normalize_sheet_name(wb, SHEET_INV_ORIG)
    
    try:
        ws_inv_orig = wb.Worksheets(normalized_inv_name)
    except Exception:
        ws_inv_orig = wb.Worksheets(1)
        normalized_inv_name = ws_inv_orig.Name

    # 4) ELIMINAR hoja INVENTARIO COPIA si existe
    try:
        excel.DisplayAlerts = False
        for i in range(1, wb.Worksheets.Count + 1):
            try:
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_COPIA):
                    wb.Worksheets(i).Delete()
                    log(f"Hoja existente eliminada: {sheet_name}")
                    break
            except:
                pass
    except Exception as e:
        log(f"Error al eliminar hoja existente: {e}")

    # 5) CREAR nueva hoja INVENTARIO COPIA
    log("Creando nueva hoja INVENTARIO COPIA...")
    was_protected = False
    try:
        if ws_inv_orig.ProtectContents or ws_inv_orig.ProtectDrawingObjects or ws_inv_orig.ProtectScenarios:
            was_protected = True
            ws_inv_orig.Unprotect(Password=PASS_INV)
    except Exception as e:
        log(f"Aviso al desproteger: {e}")

    try:
        ws_inv_copia = wb.Worksheets.Add(After=ws_inv_orig)
        ws_inv_copia.Name = SHEET_INV_COPIA
        
        ws_inv_orig.UsedRange.Copy(Destination=ws_inv_copia.Range("A1"))
        
        try:
            for col in range(1, ws_inv_orig.UsedRange.Columns.Count + 1):
                ws_inv_copia.Columns(col).ColumnWidth = ws_inv_orig.Columns(col).ColumnWidth
        except Exception as e:
            log(f"Aviso: no se pudo copiar anchos de columna: {e}")
        
        log(f"Hoja '{SHEET_INV_COPIA}' creada exitosamente")
        
        
        # ROMPER VÍNCULOS EXTERNOS: Convertir fórmulas a valores en columnas con enlaces externos
        log("Rompiendo vínculos externos en INVENTARIO COPIA...")
        try:
            # Obtener encabezados de la copia recién creada
            temp_hr, temp_hdr, temp_hdrn = ws_headers_smart(
                ws_inv_copia,
                expected_row=HEADER_ROW_INV,
                preferred_labels=["REFERENCIA", "NOMBRE LISTA", "NOMBRE MYR"]
            )
            
            # Lista de columnas que suelen tener fórmulas con enlaces externos
            columnas_a_romper = [
                "NOMBRE LISTA",      # Tiene fórmulas que apuntan a MATRIZ USD
                "NOMBRE MYR",        # Puede tener fórmulas relacionadas
                "REFERENCIA LISTA DE PRECIOS",  # Puede tener vínculos
                "PRECIO LISTA"       # Puede tener vínculos
            ]
            
            columnas_rotas = 0
            for col_name in columnas_a_romper:
                col_idx = temp_hdrn.get(_norm(col_name))
                if col_idx:
                    try:
                        # Seleccionar toda la columna desde el inicio de datos hasta el final usado
                        used_range = ws_inv_copia.UsedRange
                        last_row_temp = used_range.Rows.Count
                        
                        col_range = ws_inv_copia.Range(
                            ws_inv_copia.Cells(temp_hr + 1, col_idx),
                            ws_inv_copia.Cells(last_row_temp, col_idx)
                        )
                        
                        # Convertir fórmulas a valores
                        # Método: copiar y pegar como valores sobre sí mismo
                        col_range.Copy()
                        col_range.PasteSpecial(Paste=-4163)  # xlPasteValues
                        excel.CutCopyMode = False
                        
                        columnas_rotas += 1
                        log(f"  ✓ Vínculos rotos en columna: {col_name}")
                        
                    except Exception as e:
                        log(f"  ⚠ No se pudo romper vínculos en {col_name}: {e}")
            
            if columnas_rotas > 0:
                log(f"✓ {columnas_rotas} columna(s) convertida(s) a valores (vínculos externos eliminados)")
            else:
                log("  ℹ No se encontraron columnas con posibles vínculos externos")
                
        except Exception as e:
            log(f"  ⚠ Error al romper vínculos externos: {e}")
            # No es crítico, continuar con el proceso
        
    except Exception as e:
        log(f"ERROR al crear copia: {e}")
        raise RuntimeError(f"No se pudo crear copia de la hoja INVENTARIO: {e}")

    if was_protected:
        try:
            ws_inv_orig.Protect(Password=PASS_INV, DrawingObjects=True, Contents=True, Scenarios=True)
            log("Hoja INVENTARIO original re-protegida")
        except Exception as e:
            log(f"Aviso al re-proteger: {e}")

    # 6) TRABAJAR EN INVENTARIO COPIA
    log("Trabajando en hoja INVENTARIO COPIA...")
    
    header_row_used, hdr_copia, hdrn_copia = ws_headers_smart(
        ws_inv_copia,
        expected_row=HEADER_ROW_INV,
        preferred_labels=["REFERENCIA", "REFERENCIA FERTRAC"]
    )
    log(f"Encabezados detectados en fila {header_row_used} de INVENTARIO COPIA")

    ref_col_idx = find_reference_col_idx(hdrn_copia, ws_inv_copia, header_row_used)
    start_data_row = header_row_used + 1

    # Detectar rango inicial solo para referencia
    initial_last_row = ws_last_row(ws_inv_copia, ref_col_idx, header_row_used)
    log(f"Rango inicial detectado: {initial_last_row - start_data_row + 1} filas")

    # El last_row real se calculará después de pegar los datos
    last_row = initial_last_row

    # 7) LIMPIAR columnas en INVENTARIO COPIA
    # Calcular el rango máximo esperado ANTES de limpiar
    log("Calculando rango esperado para limpieza...")
    expected_rows = len(df_src["__REFERENCIA__"])
    max_last_row = ws_ensure_range(ws_inv_copia, start_data_row, expected_rows, header_row_used)

    log(f"Limpiando columnas en INVENTARIO COPIA (hasta fila {max_last_row})...")
    for colname in COLS_A_LIMPIAR:
        cidx = hdrn_copia.get(_norm(colname))
        if cidx:
            ws_clear_column(ws_inv_copia, cidx, start_data_row, max_last_row)


    # 8) Limpiar REFERENCIA FERTRAC en INV LISTA PRECIOS
    log("Limpiando REFERENCIA FERTRAC en INV LISTA PRECIOS...")
    try:
        ws_lp = None
        target_norm = _norm(SHEET_INV_LISTA)
        
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                ws_lp = wb.Worksheets(i)
                log(f"Hoja encontrada: '{sheet_name}'")
                break
        
        if ws_lp is None:
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name_norm = _norm(wb.Worksheets(i).Name)
                if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                    ws_lp = wb.Worksheets(i)
                    log(f"Hoja encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                    break
        
        if ws_lp:
            hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
            cidx = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
            if cidx:
                last_row_lp = ws_last_row(ws_lp, cidx, hr_lp)
                pivot_top_lp = ws_first_pivot_row(ws_lp)
                if pivot_top_lp and pivot_top_lp > hr_lp:
                    last_row_lp = min(last_row_lp, pivot_top_lp - 1)
                ws_clear_column(ws_lp, cidx, hr_lp + 1, last_row_lp)
                log("REFERENCIA FERTRAC limpiada exitosamente")
            else:
                log("Columna REFERENCIA FERTRAC no encontrada")
        else:
            log("Hoja INV LISTA PRECIOS no encontrada")
                
    except Exception as e:
        log(f"No se pudo procesar 'INV LISTA PRECIOS': {e}")

    # 9) PEGAR columnas desde datos externos en INVENTARIO COPIA
    log("Pegando columnas desde Inventario actualizado en INVENTARIO COPIA...")
    ref_values   = df_src["__REFERENCIA__"].tolist()
    nombre_odoo  = df_src.get("__NOMBRE__",       pd.Series([""]*len(ref_values))).tolist()
    marca_sys    = df_src.get("__MARCA_SYS__",    pd.Series([""]*len(ref_values))).tolist()
    linea_sys    = df_src.get("__LINEA_SYS__",    pd.Series([""]*len(ref_values))).tolist()
    sublinea_sys = df_src.get("__SUBLINEA_SYS__", pd.Series([""]*len(ref_values))).tolist()
    costo_prom   = df_src.get("__COSTO__",        pd.Series([np.nan]*len(ref_values))).tolist()

    def paste_if_exists(col_name, values, number_format=None):
        cidx = hdrn_copia.get(_norm(col_name))
        if not cidx:
            log(f"  - Columna no encontrada: {col_name}")
            return
        
        if col_name == "REFERENCIA":
            has_slash = any("/" in str(v) for v in values if v not in (None, "", np.nan))
            
            if has_slash:
                
                rng = ws_inv_copia.Range(
                    ws_inv_copia.Cells(start_data_row, cidx),
                    ws_inv_copia.Cells(start_data_row + len(values) - 1, cidx)
                )
                
                rng.NumberFormat = "@"
                ws_fill_column_values(ws_inv_copia, cidx, start_data_row, values)
                
                try:
                    converted_values = []
                    for v in values:
                        if v in (None, "", np.nan):
                            converted_values.append([""])
                        elif "/" in str(v) or not str(v).replace(".", "").replace("-", "").isdigit():
                            converted_values.append([v])
                        else:
                            try:
                                converted_values.append([float(v)])
                            except:
                                converted_values.append([v])
                    
                    rng.Value = converted_values
                except Exception as e:
                    log(f"    Aviso en conversión: {e}")
                
                rng.NumberFormat = "0"
                
                try:
                    rng.HorizontalAlignment = -4131  # xlLeft
                except Exception as e:
                    log(f"    Aviso en alineación: {e}")
                
                try:
                    for i in range(1, 8):
                        try:
                            rng.Errors.Item(i).Ignore = True
                        except:
                            pass
                    ws_inv_copia.Parent.Application.ErrorCheckingOptions.NumberAsText = False
                except Exception:
                    pass
                
                log(f"Pegada columna: {col_name} (formato número, alineación izquierda)")
                return
        
        ws_fill_column_values(ws_inv_copia, cidx, start_data_row, values)
        if number_format:
            ws_inv_copia.Columns(cidx).NumberFormat = number_format


    paste_if_exists("REFERENCIA", ref_values, number_format="0")
    paste_if_exists("NOMBRE ODOO", nombre_odoo)
    paste_if_exists("Marca sistema", marca_sys)
    paste_if_exists("Linea sistema", linea_sys)
    paste_if_exists("Sub- linea sistema", sublinea_sys)

    log("Recalculando rango de datos después de pegar...")
    new_last_row = start_data_row + len(ref_values) - 1

    # Verificar si hay pivots que limiten el rango
    pivot_top = ws_first_pivot_row(ws_inv_copia)
    if pivot_top and pivot_top > header_row_used:
        # Si los nuevos datos sobrepasan el pivot, advertir
        if new_last_row >= pivot_top:
            log(f"⚠ ADVERTENCIA: Los datos ({new_last_row} filas) sobrepasan el inicio de la tabla pivote (fila {pivot_top})")
            log(f"  Se procesarán solo las filas hasta {pivot_top - 1}")
            last_row = pivot_top - 1
        else:
            last_row = new_last_row
    else:
        last_row = new_last_row

    log(f"Rango de datos actualizado: filas {start_data_row} a {last_row} ({last_row - start_data_row + 1} registros)")


    # ELIMINAR FILAS CON "MANO DE OBRA" EN NOMBRE ODOO
    log("Eliminando filas con 'Mano de obra' en columna NOMBRE ODOO...")
    try:
        col_nombre_odoo = hdrn_copia.get(_norm("NOMBRE ODOO"))
        if col_nombre_odoo:
            # Leer valores de NOMBRE ODOO
            filas_a_eliminar = []
            for row_idx in range(start_data_row, last_row + 1):
                try:
                    valor = ws_inv_copia.Cells(row_idx, col_nombre_odoo).Value
                    if valor and isinstance(valor, str) and "mano de obra" in valor.lower():
                        filas_a_eliminar.append(row_idx)
                except Exception:
                    continue
            
            if filas_a_eliminar:
                log(f"  Encontradas {len(filas_a_eliminar)} filas con 'Mano de obra'")
                # Eliminar filas de abajo hacia arriba para mantener índices correctos
                for row_idx in reversed(filas_a_eliminar):
                    ws_inv_copia.Rows(row_idx).Delete()
                
                # Actualizar last_row
                last_row = last_row - len(filas_a_eliminar)
                log(f"  ✓ {len(filas_a_eliminar)} filas eliminadas. Nuevo rango: {start_data_row} a {last_row}")
            else:
                log("  No se encontraron filas con 'Mano de obra'")
        else:
            log("  ⚠ Columna 'NOMBRE ODOO' no encontrada")
    except Exception as e:
        log(f"  ⚠ Error al eliminar filas con 'Mano de obra': {e}")
        import traceback
        log(traceback.format_exc())


    # ELIMINAR FILAS CON REFERENCIA QUE TERMINE EN " NF" (con espacio)
    log("Eliminando filas con referencias que terminan en ' NF' (espacio + NF)...")
    try:
        # ref_col_idx ya está definido anteriormente
        if ref_col_idx:
            # Leer valores de REFERENCIA
            filas_a_eliminar = []
            for row_idx in range(start_data_row, last_row + 1):
                try:
                    valor_ref = ws_inv_copia.Cells(row_idx, ref_col_idx).Value
                    if valor_ref and isinstance(valor_ref, str):
                        # Verificar si termina con espacio + "NF"
                        if valor_ref.endswith(" NF"):
                            filas_a_eliminar.append(row_idx)
                except Exception:
                    continue
            
            if filas_a_eliminar:
                log(f"  Encontradas {len(filas_a_eliminar)} referencias terminadas en ' NF'")
                # Eliminar filas de abajo hacia arriba para mantener índices correctos
                eliminadas = 0
                for row_idx in reversed(filas_a_eliminar):
                    try:
                        ws_inv_copia.Rows(row_idx).Delete()
                        eliminadas += 1
                    except Exception as e:
                        log(f"  ⚠ Error al eliminar fila {row_idx}: {e}")
                
                # Actualizar last_row
                last_row = last_row - eliminadas
                log(f"  ✓ {eliminadas} filas eliminadas. Nuevo rango: {start_data_row} a {last_row}")
            else:
                log("  No se encontraron referencias terminadas en ' NF'")
        else:
            log("  ⚠ Columna de REFERENCIA no encontrada")
    except Exception as e:
        log(f"  ⚠ Error al eliminar referencias con ' NF': {e}")
        import traceback
        log(traceback.format_exc())

    # Actualizar el rango usado en la hoja para asegurar que Excel lo reconozca
    try:
        ws_inv_copia.UsedRange.Calculate()
    except Exception as e:
        log(f"Aviso: no se pudo recalcular UsedRange: {e}")

    # APLICAR BORDES A TODO EL RANGO
    log("Aplicando bordes a todo el rango de datos...")
    try:
        used_range = ws_inv_copia.UsedRange
        first_col = used_range.Column
        last_col = first_col + used_range.Columns.Count - 1
        ws_apply_borders_to_range(ws_inv_copia, header_row_used, last_row, first_col, last_col)
        
    except Exception as e:
        log(f"⚠ Error al aplicar bordes: {e}")
        import traceback
        log(traceback.format_exc())
    log("Limpiando formato no deseado...")
    try:
        used_range = ws_inv_copia.UsedRange
        first_col = used_range.Column
        last_col = first_col + used_range.Columns.Count - 1
        ws_remove_formatting_from_range(ws_inv_copia, start_data_row, last_row, first_col, last_col)
        
    except Exception as e:
        log(f"⚠ Error al limpiar formato: {e}")

    # ACTUALIZAR FÓRMULAS DE SUBTOTAL EN FILA 1
    log("Actualizando fórmulas de subtotal en fila 1...")
    try:
        ws_update_subtotal_formula(ws_inv_copia, 1, last_row)
    except Exception as e:
        log(f"⚠ Error al actualizar fórmulas de subtotal: {e}")

    # 10) Arrastrar fórmulas en INVENTARIO COPIA
    log("Arrastrando fórmulas en INVENTARIO COPIA...")
    for colname in ["Dif marca", "Dif linea", "Dif sub-linea"]:
        cidx = hdrn_copia.get(_norm(colname))
        if cidx:
            ws_copy_down_formula(ws_inv_copia, cidx, start_data_row, last_row)

    col_total_inv = hdrn_copia.get(_norm("TOTAL INV"))
    if col_total_inv:
        ws_copy_down_formula(ws_inv_copia, col_total_inv, start_data_row, last_row)

    col_exist = ws_ensure_existencia_header(ws_inv_copia, header_row_used)
    ws_copy_down_formula(ws_inv_copia, col_exist, start_data_row, last_row)


    # 11)Actualizar NOMBRE LISTA desde Matriz USD
    log("Actualizando NOMBRE LISTA desde Matriz USD...")
    if len(matriz_map) > 0:
        try:
            col_nombre_lista = hdrn_copia.get(_norm("NOMBRE LISTA"))
            if col_nombre_lista:
                refs_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
                refs_copia = [to_num_str(r) for r in refs_copia]
                
                descripciones = []
                matched_count = 0
                for ref in refs_copia:
                    if ref in matriz_map:
                        desc = matriz_map[ref]
                        # Si hay descripción, usarla; si no, poner "0"
                        descripciones.append(desc if desc else "0")
                        if desc:
                            matched_count += 1
                    else:
                        # Si no hay coincidencia, poner "0"
                        descripciones.append("0")
                
                write_range_as_array(ws_inv_copia, start_data_row, col_nombre_lista, descripciones)
                log(f"{matched_count} descripciones actualizadas desde Matriz USD")
            else:
                log("  ⚠ Columna 'NOMBRE LISTA' no encontrada en INVENTARIO COPIA")
        except Exception as e:
            log(f"  ⚠ Error al actualizar NOMBRE LISTA: {e}")
            import traceback
            log(traceback.format_exc())
    else:
        log("  ⚠ No hay datos de Matriz USD disponibles - saltando actualización de NOMBRE LISTA")

    # 11.5) Llenar NOMBRE MYR con prioridad NOMBRE LISTA -> NOMBRE ODOO
    log("Actualizando NOMBRE MYR (prioridad: NOMBRE LISTA → NOMBRE ODOO)...")
    try:
        col_nombre_myr = hdrn_copia.get(_norm("NOMBRE MYR"))
        col_nombre_lista = hdrn_copia.get(_norm("NOMBRE LISTA"))
        col_nombre_odoo = hdrn_copia.get(_norm("NOMBRE ODOO"))
        
        if col_nombre_myr:
            if col_nombre_lista and col_nombre_odoo:
                cols_to_read = [col_nombre_lista, col_nombre_odoo]
                data = read_multiple_columns_optimized(ws_inv_copia, start_data_row, last_row, cols_to_read)
                
                nombres_lista = data.get(col_nombre_lista, [])
                nombres_odoo = data.get(col_nombre_odoo, [])
                
                nombres_myr = []
                from_lista = 0
                from_odoo = 0
                
                for i in range(len(nombres_lista)):
                    lista_val = str(nombres_lista[i]).strip() if nombres_lista[i] not in (None, "", "None", 0) else ""
                    odoo_val = str(nombres_odoo[i]).strip() if nombres_odoo[i] not in (None, "", "None") else ""
                    
                    if lista_val:
                        nombres_myr.append(lista_val)
                        from_lista += 1
                    elif odoo_val:
                        nombres_myr.append(odoo_val)
                        from_odoo += 1
                    else:
                        nombres_myr.append("")
                
                write_range_as_array(ws_inv_copia, start_data_row, col_nombre_myr, nombres_myr)
                log(f"NOMBRE MYR actualizado: {from_lista} desde NOMBRE LISTA, {from_odoo} desde NOMBRE ODOO")
                
            elif col_nombre_lista:
                nombres_lista = read_range_as_array(ws_inv_copia, start_data_row, last_row, col_nombre_lista)
                write_range_as_array(ws_inv_copia, start_data_row, col_nombre_myr, nombres_lista)
                log(f"NOMBRE MYR copiado desde NOMBRE LISTA")
                
            elif col_nombre_odoo:
                nombres_odoo = read_range_as_array(ws_inv_copia, start_data_row, last_row, col_nombre_odoo)
                write_range_as_array(ws_inv_copia, start_data_row, col_nombre_myr, nombres_odoo)
                log(f"NOMBRE MYR copiado desde NOMBRE ODOO")
            else:
                log("  ⚠ No se encontraron columnas NOMBRE LISTA ni NOMBRE ODOO")
        else:
            log("  ⚠ Columna 'NOMBRE MYR' no encontrada en INVENTARIO COPIA")
            
    except Exception as e:
        log(f"  ⚠ Error al actualizar NOMBRE MYR: {e}")
        import traceback
        log(traceback.format_exc())

    # 12) Llevar EXISTENCIA_CALC en INVENTARIO COPIA 
    log("Escribiendo EXISTENCIA consolidada en INVENTARIO COPIA .")
    try:
        refs_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
        refs_copia = [to_num_str(r) for r in refs_copia]
        
        existencias = []
        valores_encontrados = 0
        for key in refs_copia:
            if key:
                val = exist_map.get(key)
                if pd.notna(val):
                    existencias.append(float(val))
                    valores_encontrados += 1
                else:
                    # Si no hay valor, poner 0
                    existencias.append(0)
            else:
                # Si no hay referencia, poner 0
                existencias.append(0)
        
        write_range_as_array(ws_inv_copia, start_data_row, col_exist, existencias)
        log(f"{valores_encontrados} existencias actualizadas, {len(existencias) - valores_encontrados} con valor 0")
    except Exception as e:
        log(f"⚠ Error al escribir existencias: {e}")

    # 13) Traer columnas desde INVENTARIO ORIGINAL
    log("Trayendo columnas desde INVENTARIO ORIGINAL por REFERENCIA.")
    try:
        hr_orig, hdr_orig, hdrn_orig = ws_headers_smart(ws_inv_orig, HEADER_ROW_INV, ["REFERENCIA"])
        ref_idx_orig = hdrn_orig.get(_norm("REFERENCIA")) or find_reference_col_idx(hdrn_orig, ws_inv_orig, hr_orig)
        
        if ref_idx_orig:
            last_orig = ws_last_row(ws_inv_orig, ref_idx_orig, hr_orig)
            
            pivot_top_orig = ws_first_pivot_row(ws_inv_orig)
            if pivot_top_orig and pivot_top_orig > hr_orig:
                last_orig = min(last_orig, pivot_top_orig - 1)
            
            max_rows = min(last_orig, hr_orig + 50000)
            
            log(f"Leyendo {max_rows - hr_orig} filas desde INVENTARIO ORIGINAL...")
            
            cols_to_read = {ref_idx_orig: "__REF__"}
            for colname in COLS_DESDE_ORIGINAL:
                cidx = hdrn_orig.get(_norm(colname))
                if cidx:
                    cols_to_read[cidx] = colname
            
            if len(cols_to_read) <= 1:
                log("⚠ No hay columnas adicionales para traer")
            else:
                col_indices = sorted(cols_to_read.keys())
                all_data = read_multiple_columns_optimized(ws_inv_orig, hr_orig + 1, max_rows, col_indices)
                
                refs_orig = all_data[ref_idx_orig]
                refs_orig_normalized = [to_num_str(r) for r in refs_orig]
                
                # OPTIMIZADO: Solo leer texto de CLASIFICACION (donde están los #N/D)
                text_data = {}
                clasificacion_idx = None
                
                for col_idx in col_indices:
                    if col_idx == ref_idx_orig:
                        continue
                    colname = cols_to_read[col_idx]
                    
                    # Solo procesar CLASIFICACION para detectar #N/D
                    if _norm(colname) == _norm("CLASIFICACION"):
                        clasificacion_idx = col_idx
                        log(f"Detectando errores #N/D en columna {colname}...")
                        
                        try:
                            # Leer como array 2D (más rápido que iterar)
                            rng = ws_inv_orig.Range(
                                ws_inv_orig.Cells(hr_orig + 1, col_idx),
                                ws_inv_orig.Cells(max_rows, col_idx)
                            )
                            
                            # Obtener valores como array
                            arr = rng.Value
                            if arr is None:
                                text_values = [None] * (max_rows - hr_orig)
                            elif isinstance(arr, (list, tuple)):
                                text_values = []
                                for row in arr:
                                    if isinstance(row, (list, tuple)):
                                        val = row[0] if row else None
                                    else:
                                        val = row
                                    
                                    # Verificar si es un error de Excel (número negativo específico)
                                    # En COM, #N/D se representa como -2146826246
                                    if val == -2146826246 or (isinstance(val, str) and val.startswith("#")):
                                        text_values.append("#N/D")
                                    elif isinstance(val, int) and val < 0 and val > -2147000000:
                                        # Otros errores de Excel
                                        text_values.append(f"#{val}")
                                    else:
                                        text_values.append(None)
                            else:
                                text_values = [None]
                            
                            text_data[colname] = text_values
                            errores_detectados = sum(1 for v in text_values if v is not None)
                            if errores_detectados > 0:
                                log(f"{errores_detectados} errores #N/D detectados en {colname}")
                            
                        except Exception as e:
                            log(f"  ⚠ Error al leer texto de {colname}: {e}")
                            text_data[colname] = [None] * (max_rows - hr_orig)
                
                maps = {}
                maps_text = {}
                
                for col_idx in col_indices:
                    if col_idx == ref_idx_orig:
                        continue
                    colname = cols_to_read[col_idx]
                    maps[colname] = dict(zip(refs_orig_normalized, all_data[col_idx]))
                    
                    # Crear mapa de textos con errores solo para CLASIFICACION
                    if colname in text_data:
                        maps_text[colname] = {}
                        for idx, ref in enumerate(refs_orig_normalized):
                            if idx < len(text_data[colname]) and text_data[colname][idx]:
                                maps_text[colname][ref] = text_data[colname][idx]
                
                log("Leyendo referencias de INVENTARIO COPIA...")
                refs_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
                refs_copia_normalized = [to_num_str(r) for r in refs_copia]
                
                log("Construyendo valores a escribir...")
                for colname in COLS_DESDE_ORIGINAL:
                    tgt_idx = hdrn_copia.get(_norm(colname))
                    if not tgt_idx or colname not in maps:
                        continue
                    
                    values_to_write = []
                    matched = 0
                    errores_preservados = 0
                    
                    for ref in refs_copia_normalized:
                        if ref and ref in maps[colname]:
                            # Verificar si hay un error de Excel (#N/D) en esta columna
                            if colname in maps_text and ref in maps_text[colname]:
                                texto_error = maps_text[colname][ref]
                                values_to_write.append(texto_error)
                                matched += 1
                                errores_preservados += 1
                            else:
                                val = maps[colname][ref]
                                values_to_write.append(val if val not in (None, "", "None") else "")
                                if val not in (None, "", "None"):
                                    matched += 1
                        else:
                            values_to_write.append("")
                    
                    write_range_as_array(ws_inv_copia, start_data_row, tgt_idx, values_to_write)
                    
                    if errores_preservados > 0:
                        log(f"  {colname}: {matched} valores ({errores_preservados} errores #N/D preservados)")
                    else:
                        log(f"  {colname}: {matched} valores copiados")
                
                log(f"Columnas traídas exitosamente desde INVENTARIO ORIGINAL")
                
                # Forzar formato de texto solo en CLASIFICACION con #N/D
                try:
                    if "CLASIFICACION" in maps_text and maps_text["CLASIFICACION"]:
                        tgt_idx = hdrn_copia.get(_norm("CLASIFICACION"))
                        if tgt_idx:
                            log("Aplicando formato de texto a celdas con #N/D...")
                            for idx, ref in enumerate(refs_copia_normalized):
                                if ref in maps_text["CLASIFICACION"]:
                                    fila = start_data_row + idx
                                    cell = ws_inv_copia.Cells(fila, tgt_idx)
                                    cell.NumberFormat = "@"
                                    cell.Value = maps_text["CLASIFICACION"][ref]
                            log(f"Formato aplicado a {len(maps_text['CLASIFICACION'])} celdas")
                            
                except Exception as e:
                    log(f"  ⚠ Error al forzar formato de texto: {e}")
                
                try:
                    inv_bodega_idx = hdrn_copia.get(_norm("INV BODEGA GERENCIA"))
                    if inv_bodega_idx:
                        rng = ws_inv_copia.Range(
                            ws_inv_copia.Cells(start_data_row, inv_bodega_idx),
                            ws_inv_copia.Cells(last_row, inv_bodega_idx)
                        )
                        rng.HorizontalAlignment = -4108  # xlCenter
                    else:
                        log("  ⚠ Columna INV BODEGA GERENCIA no encontrada")
                except Exception as e:
                    log(f"  ⚠ Error al centrar INV BODEGA GERENCIA: {e}")
                
    except Exception as e:
        log(f"⚠ Error al traer columnas desde original: {e}")
        import traceback
        log(traceback.format_exc())


    # AGREGAR SUBTOTALES FINALES 
    log("Agregando subtotales finales...")
    try:
        ws_add_final_subtotals(ws_inv_copia, last_row, header_row_used, hdrn_copia)
    except Exception as e:
        log(f"⚠ Error al agregar subtotales finales: {e}")

    try:
        # PRIMERO: Activar la hoja
        ws_inv_copia.Activate()
        
        # SEGUNDO: Seleccionar la celda A3
        ws_inv_copia.Cells(3, 1).Select()
        
        # TERCERO: Aplicar FreezePanes
        excel.ActiveWindow.FreezePanes = True
        log("✓ Paneles inmovilizados en fila 3")
    except Exception as e:
        log(f"⚠ Error al inmovilizar paneles: {e}")

    # NUEVO: Eliminar registros según criterios estandarizados
    log("")
    log("FASE: Eliminación de registros estandarizados")
    last_row = eliminar_registros_estandarizados(
        ws_inv_copia, 
        start_data_row, 
        last_row, 
        ref_col_idx, 
        hdrn_copia
    )
        
    # Aplicar reglas de marcas propias 
    log("Aplicando reglas de negocio para marcas propias...")
    last_row = aplicar_reglas_marcas_propias(
        ws_inv_copia, 
        start_data_row, 
        last_row, 
        ref_col_idx, 
        hdrn_copia, 
        marcas_propias, 
        distribucion
    )

    # #Eliminar registros con LINEA COPIA indeterminada
    # last_row = eliminar_registros_linea_copia_indeterminada(
    #     ws_inv_copia, 
    #     start_data_row, 
    #     last_row, 
    #     ref_col_idx, 
    #     hdrn_copia
    # )

    # ELIMINAR FILAS CON LIDER LINEA VACÍO

    # log("ELIMINANDO FILAS CON LIDER LINEA VACÍO")
    # try:
    #     col_lider_linea = hdrn_copia.get(_norm("LIDER LINEA"))
        
    #     if col_lider_linea:
    #         # Actualizar last_row antes de procesar
    #         last_row = ws_last_row(ws_inv_copia, ref_col_idx, start_data_row - 1)
    #         log(f"Analizando {last_row - start_data_row + 1} filas...")
            
    #         # Leer valores de LIDER LINEA
    #         filas_a_eliminar = []
    #         for row_idx in range(start_data_row, last_row + 1):
    #             try:
    #                 valor_lider = ws_inv_copia.Cells(row_idx, col_lider_linea).Value
    #                 # Verificar si está vacío (None, "", espacios, etc.)
    #                 if not valor_lider or (isinstance(valor_lider, str) and not valor_lider.strip()):
    #                     filas_a_eliminar.append(row_idx)
    #             except Exception:
    #                 continue
            
    #         if filas_a_eliminar:
    #             log(f"  Encontradas {len(filas_a_eliminar)} filas con LIDER LINEA vacío")
                
    #             # Eliminar filas de abajo hacia arriba para mantener índices correctos
    #             eliminadas = 0
    #             for row_idx in reversed(filas_a_eliminar):
    #                 try:
    #                     ws_inv_copia.Rows(row_idx).Delete()
    #                     eliminadas += 1
    #                 except Exception as e:
    #                     log(f"  ⚠ Error al eliminar fila {row_idx}: {e}")
                
    #             # Actualizar last_row
    #             last_row = last_row - eliminadas
    #             log(f"  ✓ {eliminadas} filas eliminadas. Nuevo rango: {start_data_row} a {last_row} ({last_row - start_data_row + 1} registros)")
    #         else:
    #             log("  ℹ No se encontraron filas con LIDER LINEA vacío")
    #     else:
    #         log("  ⚠ Columna 'LIDER LINEA' no encontrada")
            
    # except Exception as e:
    #     log(f"❌ ERROR al eliminar filas con LIDER LINEA vacío: {e}")
    #     import traceback
    #     log(traceback.format_exc())
    
    # log("")
    
    # last_row = procesar_existencias_negativas_y_cero(
    #     ws_inv_copia,
    #     start_data_row,
    #     last_row,
    #     ref_col_idx,
    #     hdrn_copia,
    #     BASE_PATH
    # )


    # LLENAR COSTO PROMEDIO 
    log("Actualizando COSTO PROMEDIO después de calcular existencias...")
    try:
        col_costo_promedio = hdrn_copia.get(_norm("COSTO PROMEDIO"))
        
        if not col_costo_promedio:
            log("  ⚠ Columna COSTO PROMEDIO no encontrada en INVENTARIO COPIA")
        else:
            # Verificar que tenemos los datos de costo desde el archivo fuente
            if "__COSTO__" in df_src.columns:
                # Leer referencias actuales de INVENTARIO COPIA
                refs_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
                refs_copia_norm = [to_num_str(r) for r in refs_copia]
                
                # Leer EXISTENCIAS actuales para aplicar la regla
                existencias_actuales = read_range_as_array(ws_inv_copia, start_data_row, last_row, col_exist)
                
                # Crear mapa de REFERENCIA -> COSTO desde el archivo fuente
                costo_map = dict(zip(
                    df_src["__REFERENCIA__"].apply(to_num_str),
                    df_src["__COSTO__"]
                ))
                
                # Cruzar y llenar COSTO PROMEDIO con regla: si existencia = 0, entonces costo = 0
                costos = []
                matched = 0
                costos_cero_por_existencia = 0
                
                for i, ref in enumerate(refs_copia_norm):
                    # Obtener existencia para esta fila
                    try:
                        exist_val = float(existencias_actuales[i]) if existencias_actuales[i] not in (None, "", "None") else 0.0
                    except:
                        exist_val = 0.0
                    
                    # REGLA: Si existencia es 0, costo es 0
                    if exist_val == 0:
                        costos.append(0)
                        costos_cero_por_existencia += 1
                    elif ref and ref in costo_map:
                        costo_val = costo_map[ref]
                        
                        # Validar que sea un valor numérico válido
                        if pd.notna(costo_val):
                            try:
                                costos.append(float(costo_val))
                                matched += 1
                            except:
                                costos.append("")
                        else:
                            costos.append("")
                    else:
                        costos.append("")
                
                # Escribir los valores de costo
                write_range_as_array(ws_inv_copia, start_data_row, col_costo_promedio, costos)
                
                log(f"COSTO PROMEDIO actualizado:")

                # Aplicar formato numérico de contabilidad (opcional)
                try:
                    rng = ws_inv_copia.Range(
                        ws_inv_copia.Cells(start_data_row, col_costo_promedio),
                        ws_inv_copia.Cells(last_row, col_costo_promedio)
                    )
                    # Formato contabilidad sin decimales, con separador de miles (punto)
                    rng.NumberFormat = "_($* #.##0_);_($* (#.##0);_($* \"-\"_);_(@_)"
                    log(f"   - Formato de contabilidad aplicado (sin decimales, con separador de miles)")
                except Exception as e:
                    log(f"   ⚠ No se pudo aplicar formato: {e}")

                    
            else:
                log("  ⚠ No hay datos de COSTO en el archivo fuente (columna __COSTO__)")
                
    except Exception as e:
        log(f"❌ ERROR al actualizar COSTO PROMEDIO: {e}")
        import traceback
        log(traceback.format_exc())

    # 14) Llenar REFERENCIA FERTRAC en INV LISTA PRECIOS
    log("Llenando REFERENCIA FERTRAC en INV LISTA PRECIOS desde INVENTARIO COPIA...")
    try:
        ws_lp = None
        target_norm = _norm(SHEET_INV_LISTA)
        
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                ws_lp = wb.Worksheets(i)
                break
        
        if ws_lp is None:
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name_norm = _norm(wb.Worksheets(i).Name)
                if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                    ws_lp = wb.Worksheets(i)
                    break
        
        if ws_lp:
            hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
            ref_fertrac_idx = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
            
            if ref_fertrac_idx:
                log(f"✓ Columna REFERENCIA FERTRAC encontrada en índice {ref_fertrac_idx}")
                
                # Leer referencias desde INVENTARIO COPIA
                referencias_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
                
                # 🔥 LIMPIAR referencias (eliminar .0, notación científica, etc.)
                referencias_copia = [limpiar_referencia(r) for r in referencias_copia if limpiar_referencia(r)]
                
                log(f"📋 {len(referencias_copia)} referencias a copiar")
                
                # Calcular last_row basándose en el número de referencias
                last_row_lp = hr_lp + len(referencias_copia)
                
                # 🔥 APLICAR FORMATO DE TEXTO ANTES DE ESCRIBIR
                log("⚙️  Aplicando formato de TEXTO a la columna...")
                rng = ws_lp.Range(
                    ws_lp.Cells(hr_lp + 1, ref_fertrac_idx),
                    ws_lp.Cells(last_row_lp, ref_fertrac_idx)
                )
                rng.NumberFormat = "@"  # Formato TEXTO
                
                # Escribir referencias
                write_range_as_array(ws_lp, hr_lp + 1, ref_fertrac_idx, referencias_copia)
                
                # 🔥 RE-APLICAR formato de texto después de escribir (por seguridad)
                rng.NumberFormat = "@"
                rng.HorizontalAlignment = -4131  # xlLeft (alineación izquierda)
                
                # Ignorar advertencias de "número almacenado como texto"
                try:
                    for i in range(1, 8):
                        try:
                            rng.Errors.Item(i).Ignore = True
                        except:
                            pass
                    ws_lp.Parent.Application.ErrorCheckingOptions.NumberAsText = False
                    log(f"✓ Advertencias de Excel desactivadas")
                except Exception as e:
                    log(f"   ⚠️  No se pudieron desactivar advertencias: {e}")
                
                log(f"✅ {len(referencias_copia)} referencias copiadas con formato TEXTO")
                
            else:
                log("  ⚠️  No se encontró columna REFERENCIA FERTRAC")
        else:
            log("  ⚠️  No se encontró la hoja INV LISTA PRECIOS")
            
    except Exception as e:
        log(f"  ❌ ERROR al llenar REFERENCIA FERTRAC: {e}")
        import traceback
        log(traceback.format_exc())

    # 15) Llenar REFERENCIA LISTA DE PRECIOS en INV LISTA PRECIOS desde MATRIZ USD
    log("Llenando REFERENCIA LISTA DE PRECIOS desde MATRIZ USD...")
    try:
        # Verificar que tenemos datos de Matriz USD
        if len(matriz_map_ref_lista) == 0:
            log("  ⚠ No hay datos de REFERENCIA LISTA DE PRECIOS en Matriz USD - saltando")
        else:
            # Buscar la hoja INV LISTA PRECIOS
            ws_lp = None
            target_norm = _norm(SHEET_INV_LISTA)
            
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                    ws_lp = wb.Worksheets(i)
                    log(f"✓ Hoja encontrada: '{sheet_name}'")
                    break
            
            if ws_lp is None:
                for i in range(1, wb.Worksheets.Count + 1):
                    sheet_name_norm = _norm(wb.Worksheets(i).Name)
                    if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                        ws_lp = wb.Worksheets(i)
                        log(f"✓ Hoja encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                        break
            
            if ws_lp:
                # Obtener encabezados de INV LISTA PRECIOS
                hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
                
                # Buscar columnas necesarias
                ref_fertrac_idx = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
                ref_lista_idx = hdrn_lp.get(_norm("REFERENCIA LISTA DE PRECIOS")) or \
                            hdrn_lp.get(_norm("REFERENCIA LISTA")) or \
                            hdrn_lp.get(_norm("REF LISTA PRECIOS"))
                
                if not ref_fertrac_idx:
                    log("  ⚠ Columna REFERENCIA FERTRAC no encontrada en INV LISTA PRECIOS")
                elif not ref_lista_idx:
                    log("  ⚠ Columna REFERENCIA LISTA DE PRECIOS no encontrada en INV LISTA PRECIOS")
                    log(f"     Columnas disponibles: {list(hdr_lp.keys())}")
                else:
                    log(f"✓ Columnas encontradas:")
                    log(f"  - REFERENCIA FERTRAC: índice {ref_fertrac_idx}")
                    log(f"  - REFERENCIA LISTA DE PRECIOS: índice {ref_lista_idx}")
                    
                    # Determinar última fila con datos (después del paso 14)
                    last_row_lp = ws_last_row(ws_lp, ref_fertrac_idx, hr_lp)
                    pivot_top_lp = ws_first_pivot_row(ws_lp)
                    if pivot_top_lp and pivot_top_lp > hr_lp:
                        last_row_lp = min(last_row_lp, pivot_top_lp - 1)
                    
                    log(f"📋 Procesando {last_row_lp - hr_lp} filas...")
                    
                    # Leer REFERENCIA FERTRAC de INV LISTA PRECIOS
                    refs_fertrac_lp = read_range_as_array(ws_lp, hr_lp + 1, last_row_lp, ref_fertrac_idx)
                    refs_fertrac_lp_norm = [to_num_str(r) for r in refs_fertrac_lp]
                    
                    # Cruzar con MATRIZ USD para obtener REFERENCIA LISTA DE PRECIOS
                    refs_lista_precios = []
                    matched = 0

                    for ref_fertrac in refs_fertrac_lp_norm:
                        if ref_fertrac and ref_fertrac in matriz_map_ref_lista:
                            ref_lista_val = matriz_map_ref_lista[ref_fertrac]
                            
                            # 🔥 LIMPIAR referencia antes de agregar
                            val_limpio = limpiar_referencia(ref_lista_val)
                            
                            if val_limpio:
                                refs_lista_precios.append(val_limpio)
                                matched += 1
                            else:
                                refs_lista_precios.append("0")
                        else:
                            refs_lista_precios.append("0")
                    
                    # 🔥 APLICAR FORMATO DE TEXTO ANTES DE ESCRIBIR
                    log("⚙️  Aplicando formato de TEXTO a REFERENCIA LISTA DE PRECIOS...")
                    last_row_ref_lista = hr_lp + len(refs_lista_precios)
                    
                    rng = ws_lp.Range(
                        ws_lp.Cells(hr_lp + 1, ref_lista_idx),
                        ws_lp.Cells(last_row_ref_lista, ref_lista_idx)
                    )
                    rng.NumberFormat = "@"  # Formato TEXTO
                    
                    # Escribir valores
                    write_range_as_array(ws_lp, hr_lp + 1, ref_lista_idx, refs_lista_precios)
                    
                    # 🔥 RE-APLICAR formato de texto después de escribir
                    rng.NumberFormat = "@"
                    rng.HorizontalAlignment = -4131  # xlLeft
                    
                    # Ignorar advertencias de "número almacenado como texto"
                    try:
                        for i in range(1, 8):
                            try:
                                rng.Errors.Item(i).Ignore = True
                            except:
                                pass
                        ws_lp.Parent.Application.ErrorCheckingOptions.NumberAsText = False
                        log(f"✓ Advertencias de Excel desactivadas para REFERENCIA LISTA DE PRECIOS")
                    except Exception as e:
                        log(f"   ⚠️  No se pudieron desactivar advertencias: {e}")
                    
                    log(f"✅ REFERENCIA LISTA DE PRECIOS actualizada:")
                    log(f"   - Total procesado: {len(refs_lista_precios)}")
                    log(f"   - Coincidencias encontradas: {matched}")
                    log(f"   - Sin coincidencia: {len(refs_lista_precios) - matched}")
                    
            else:
                log("  ⚠ No se encontró la hoja INV LISTA PRECIOS")
                
    except Exception as e:
        log(f"  ❌ ERROR al llenar REFERENCIA LISTA DE PRECIOS: {e}")
        import traceback
        log(traceback.format_exc())   
      
    # 16) Llenar EXISTENCIA (con fecha) en INV LISTA PRECIOS desde INVENTARIO COPIA
    log("Llenando EXISTENCIA (con fecha) en INV LISTA PRECIOS...")
    try:
        # Buscar la hoja INV LISTA PRECIOS
        ws_lp = None
        target_norm = _norm(SHEET_INV_LISTA)
        
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                ws_lp = wb.Worksheets(i)
                log(f"Hoja encontrada: '{sheet_name}'")
                break
        
        if ws_lp is None:
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name_norm = _norm(wb.Worksheets(i).Name)
                if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                    ws_lp = wb.Worksheets(i)
                    log(f"Hoja encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                    break
        
        if ws_lp:
            # Obtener encabezados de INV LISTA PRECIOS
            hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
            
            # Buscar columna REFERENCIA FERTRAC en INV LISTA PRECIOS
            ref_fertrac_idx_lp = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
            
            # Buscar columna EXISTENCIA (con fecha) en INV LISTA PRECIOS
            # Buscar cualquier columna que empiece con "EXISTENCIA"
            exist_col_lp = None
            for name, col in hdr_lp.items():
                if _norm(name).startswith("existencia "):
                    exist_col_lp = col
                    log(f"Columna EXISTENCIA encontrada en INV LISTA PRECIOS: '{name}' (índice {col})")
                    break
            
            if not ref_fertrac_idx_lp:
                log("  ⚠ Columna REFERENCIA FERTRAC no encontrada en INV LISTA PRECIOS")
            elif not exist_col_lp:
                log("  ⚠ Columna EXISTENCIA no encontrada en INV LISTA PRECIOS")
                log(f"     Columnas disponibles: {list(hdr_lp.keys())}")
            else:
                # Actualizar el encabezado con la fecha actual
                target_header = exist_col_title_for_today()
                ws_lp.Cells(hr_lp, exist_col_lp).Value = target_header
                log(f"Encabezado actualizado a: '{target_header}'")
                
                # Determinar última fila con datos en INV LISTA PRECIOS
                last_row_lp = ws_last_row(ws_lp, ref_fertrac_idx_lp, hr_lp)
                pivot_top_lp = ws_first_pivot_row(ws_lp)
                if pivot_top_lp and pivot_top_lp > hr_lp:
                    last_row_lp = min(last_row_lp, pivot_top_lp - 1)
                
                log(f"  Procesando {last_row_lp - hr_lp} filas...")
                
                # Buscar columna EXISTENCIA en INVENTARIO COPIA
                exist_col_inv_copia = None
                for name, col in hdrn_copia.items():
                    if name.startswith(_norm("EXISTENCIA")):
                        exist_col_inv_copia = col
                        break
                
                if not exist_col_inv_copia:
                    log("  ⚠ Columna EXISTENCIA no encontrada en INVENTARIO COPIA")
                else:
                    log(f"Columna EXISTENCIA encontrada en INVENTARIO COPIA: índice {exist_col_inv_copia}")
                    
                    # Leer REFERENCIA FERTRAC de INV LISTA PRECIOS
                    refs_fertrac_lp = read_range_as_array(ws_lp, hr_lp + 1, last_row_lp, ref_fertrac_idx_lp)
                    refs_fertrac_lp_norm = [to_num_str(r) for r in refs_fertrac_lp]
                    
                    # Leer REFERENCIA y EXISTENCIA de INVENTARIO COPIA
                    refs_inv_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
                    refs_inv_copia_norm = [to_num_str(r) for r in refs_inv_copia]
                    
                    existencias_inv_copia = read_range_as_array(ws_inv_copia, start_data_row, last_row, exist_col_inv_copia)
                    
                    # Crear diccionario de REFERENCIA -> EXISTENCIA desde INVENTARIO COPIA
                    exist_map_inv = dict(zip(refs_inv_copia_norm, existencias_inv_copia))
                    
                    # Cruzar y llenar EXISTENCIA en INV LISTA PRECIOS
                    existencias_lp = []
                    matched = 0
                    
                    for ref_fertrac in refs_fertrac_lp_norm:
                        if ref_fertrac and ref_fertrac in exist_map_inv:
                            exist_val = exist_map_inv[ref_fertrac]
                            
                            # Convertir a número si es posible
                            try:
                                if exist_val is not None and exist_val not in ("", "None"):
                                    exist_num = float(exist_val)
                                    existencias_lp.append(exist_num)
                                    matched += 1
                                else:
                                    existencias_lp.append(0)
                            except:
                                existencias_lp.append(0)
                        else:
                            # No hay coincidencia
                            existencias_lp.append(0)
                    
                    # Escribir en EXISTENCIA de INV LISTA PRECIOS
                    write_range_as_array(ws_lp, hr_lp + 1, exist_col_lp, existencias_lp)
                    
                    log(f"EXISTENCIA actualizada en INV LISTA PRECIOS:")
                    log(f" - Total procesado: {len(existencias_lp)}")
                    log(f" - Coincidencias encontradas: {matched}")
                    log(f" - Sin coincidencia (valor 0): {len(existencias_lp) - matched}")
                    
        else:
            log("  ⚠ No se encontró la hoja INV LISTA PRECIOS")
            
    except Exception as e:
        log(f"  ❌ ERROR al llenar EXISTENCIA en INV LISTA PRECIOS: {e}")
        import traceback
        log(traceback.format_exc())
    

    # 17) Llenar UND REM CONSIGNACION en INV LISTA PRECIOS desde Mayor Existencia
    log("Llenando UND REM CONSIGNACION en INV LISTA PRECIOS...")
    try:
        # Verificar que tenemos datos de Mayor Existencia
        if len(mayor_exist_map) == 0:
            log("  ⚠ No hay datos de Mayor Existencia disponibles - saltando")
        else:
            # Buscar la hoja INV LISTA PRECIOS
            ws_lp = None
            target_norm = _norm(SHEET_INV_LISTA)
            
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                    ws_lp = wb.Worksheets(i)
                    log(f"Hoja encontrada: '{sheet_name}'")
                    break
            
            if ws_lp is None:
                for i in range(1, wb.Worksheets.Count + 1):
                    sheet_name_norm = _norm(wb.Worksheets(i).Name)
                    if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                        ws_lp = wb.Worksheets(i)
                        log(f"Hoja encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                        break
            
            if ws_lp:
                # Obtener encabezados de INV LISTA PRECIOS
                hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
                
                # Buscar columnas necesarias
                ref_fertrac_idx_lp = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
                
                # Buscar columna UND REM CONSIGNACION (exacta y variantes)
                rem_consig_idx = (
                    hdrn_lp.get(_norm("UND REM CONSIGNACION")) or
                    hdrn_lp.get(_norm("UND REM CONSIG")) or
                    hdrn_lp.get(_norm("REM CONSIGNACION")) or
                    hdrn_lp.get(_norm("REM EN CONSIGNACION"))
                )
                
                if not ref_fertrac_idx_lp:
                    log("  ⚠ Columna REFERENCIA FERTRAC no encontrada en INV LISTA PRECIOS")
                elif not rem_consig_idx:
                    log("  ⚠ Columna UND REM CONSIGNACION no encontrada en INV LISTA PRECIOS")
                    log(f"     Columnas disponibles: {list(hdr_lp.keys())}")
                    # Mostrar columnas que contengan "consig" o "rem"
                    posibles = [k for k in hdr_lp.keys() if 'consig' in _norm(k) or 'rem' in _norm(k)]
                    if posibles:
                        log(f"     Columnas posibles con 'consig' o 'rem': {posibles}")
                else:
                    # Obtener el nombre real de la columna para el log
                    col_name_real = [k for k, v in hdr_lp.items() if v == rem_consig_idx][0]
                    
                    log(f"Columnas encontradas:")
                    log(f" - REFERENCIA FERTRAC: índice {ref_fertrac_idx_lp}")
                    log(f" - UND REM CONSIGNACION: '{col_name_real}' (índice {rem_consig_idx})")
                    
                    # Determinar última fila con datos en INV LISTA PRECIOS
                    last_row_lp = ws_last_row(ws_lp, ref_fertrac_idx_lp, hr_lp)
                    pivot_top_lp = ws_first_pivot_row(ws_lp)
                    if pivot_top_lp and pivot_top_lp > hr_lp:
                        last_row_lp = min(last_row_lp, pivot_top_lp - 1)
                    
                    log(f"   Procesando {last_row_lp - hr_lp} filas...")
                    
                    # Leer REFERENCIA FERTRAC de INV LISTA PRECIOS
                    refs_fertrac_lp = read_range_as_array(ws_lp, hr_lp + 1, last_row_lp, ref_fertrac_idx_lp)
                    refs_fertrac_lp_norm = [to_num_str(r) for r in refs_fertrac_lp]
                    
                    # Cruzar con Mayor Existencia (REM EN CONSIG) para llenar UND REM CONSIGNACION
                    valores_rem_consig = []
                    matched = 0
                    valores_no_cero = 0
                    
                    for ref_fertrac in refs_fertrac_lp_norm:
                        if ref_fertrac and ref_fertrac in mayor_exist_map:
                            val = mayor_exist_map[ref_fertrac]
                            
                            # Convertir a número
                            try:
                                val_num = float(val) if val is not None else ""  
                                
                                # Si el valor es 0 desde la fuente, sí lo ponemos
                                if val_num == 0 or val_num == "":
                                    valores_rem_consig.append("" if val is None or val == "" else 0)
                                else:
                                    valores_rem_consig.append(val_num)
                                    matched += 1
                                    valores_no_cero += 1
                            except:
                                valores_rem_consig.append("")  
                        else:
                            # No hay coincidencia - dejar en blanco
                            valores_rem_consig.append("")  

                    # Escribir en UND REM CONSIGNACION de INV LISTA PRECIOS
                    write_range_as_array(ws_lp, hr_lp + 1, rem_consig_idx, valores_rem_consig)
                    
                    log(f"UND REM CONSIGNACION actualizada en INV LISTA PRECIOS:")
                    log(f" - Total procesado: {len(valores_rem_consig)}")
                    log(f" - Coincidencias encontradas: {matched}")
                    log(f" - Valores diferentes de cero: {valores_no_cero}")
                    log(f" - Sin coincidencia (valor 0): {len(valores_rem_consig) - matched}")
                    
            else:
                log("  ⚠ No se encontró la hoja INV LISTA PRECIOS")
                
    except Exception as e:
        log(f"  ❌ ERROR al llenar UND REM CONSIGNACION: {e}")
        import traceback
        log(traceback.format_exc())


    # 17.5) APLICAR BORDES A INV LISTA PRECIOS
    log("Aplicando bordes a INV LISTA PRECIOS...")
    try:
        # Buscar la hoja INV LISTA PRECIOS
        ws_lp = None
        target_norm = _norm(SHEET_INV_LISTA)
        
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                ws_lp = wb.Worksheets(i)
                log(f"  Hoja encontrada: '{sheet_name}'")
                break
        
        if ws_lp is None:
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name_norm = _norm(wb.Worksheets(i).Name)
                if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                    ws_lp = wb.Worksheets(i)
                    log(f"  Hoja encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                    break
        
        if ws_lp:
            # Obtener encabezados
            hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
            
            # Buscar columna de referencia para determinar última fila con datos
            ref_col_lp = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
            
            if ref_col_lp:
                # Determinar última fila con datos
                last_row_lp = ws_last_row(ws_lp, ref_col_lp, hr_lp)
                
                # Ajustar por pivots si existen
                pivot_top_lp = ws_first_pivot_row(ws_lp)
                if pivot_top_lp and pivot_top_lp > hr_lp:
                    last_row_lp = min(last_row_lp, pivot_top_lp - 1)
                
                # Determinar rango de columnas
                used_range_lp = ws_lp.UsedRange
                first_col_lp = used_range_lp.Column
                last_col_lp = first_col_lp + used_range_lp.Columns.Count - 1
                
                log(f"  Aplicando bordes: filas {hr_lp} a {last_row_lp}, columnas {first_col_lp} a {last_col_lp}")
                
                # Aplicar bordes a todo el rango
                ws_apply_borders_to_range(ws_lp, hr_lp, last_row_lp, first_col_lp, last_col_lp)
                
            else:
                log("  ⚠ No se pudo determinar columna de referencia para aplicar bordes")
        else:
            log("  ⚠ No se encontró la hoja INV LISTA PRECIOS para aplicar bordes")
            
    except Exception as e:
        log(f"  ❌ ERROR al aplicar bordes a INV LISTA PRECIOS: {e}")
        import traceback
        log(traceback.format_exc())

    
    # 18) GUARDADO COMO ARCHIVO NUEVO 
    log("Preparando guardado del archivo...")

    try:
        ws_count = int(wb.Worksheets.Count)
        has_visible = False
        for i in range(1, ws_count + 1):
            try:
                if int(wb.Worksheets(i).Visible) == -1:
                    has_visible = True
                    break
            except Exception:
                pass
        if not has_visible and ws_count >= 1:
            wb.Worksheets(1).Visible = -1
            wb.Worksheets(1).Activate()
    except Exception:
        pass

    with contextlib.suppress(Exception):
        wb.IsAddin = False
    with contextlib.suppress(Exception):
        wb.Windows(1).Visible = True

    try:
        base_name = OUTPUT_BASENAME
    except NameError:
        base_name = f"{Path(PATRON_INV_FILE).stem} (ACTUALIZADO)"
    out_name = f"{base_name} {datetime.now():%Y%m%d_%H%M}.xlsx"
    out_path = BASE_PATH / out_name

    log(f"Guardando archivo (sin ordenar): {out_name}")
    apply_pw = saveinfo.get("reapply_password")
    if apply_pw:
        wb.SaveAs(str(out_path), FileFormat=51, Password=apply_pw)
    else:
        wb.SaveAs(str(out_path), FileFormat=51)

  
    # Aplicar ordenamiento DESPUÉS de guardar
    log("Aplicando autofiltros y ordenamiento por TOTAL INV...")
    try:
        # Activar la hoja INVENTARIO COPIA
        ws_inv_copia.Activate()
        
        aplicar_autofiltros_y_ordenar(ws_inv_copia, header_row_used, last_row, hdrn_copia)
        
        # Restaurar cálculo automático AHORA
        try:
            excel.Calculation = -4105  
        except Exception as e:
            log(f"Aviso al restaurar cálculo: {e}")
        
        # GUARDAR DE NUEVO con el ordenamiento aplicado
        wb.Save()

        
    except Exception as e:
        log(f"⚠ Error al aplicar ordenamiento: {e}")
        import traceback
        log(traceback.format_exc())

    # ORDENAR INV LISTA PRECIOS SEGÚN ORDEN FINAL DE INVENTARIO COPIA
    log("Ordenando INV LISTA PRECIOS según orden FINAL de INVENTARIO COPIA (después de ordenar por TOTAL INV)...")
    try:
        # Buscar la hoja INV LISTA PRECIOS
        ws_lp = None
        target_norm = _norm(SHEET_INV_LISTA)
        
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                ws_lp = wb.Worksheets(i)
                log(f"  Hoja INV LISTA PRECIOS encontrada: '{sheet_name}'")
                break
        
        if ws_lp is None:
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name_norm = _norm(wb.Worksheets(i).Name)
                if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                    ws_lp = wb.Worksheets(i)
                    log(f"  Hoja INV LISTA PRECIOS encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                    break
        
        if ws_lp:
            # Obtener encabezados de INV LISTA PRECIOS
            hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
            
            # Buscar columna REFERENCIA FERTRAC en INV LISTA PRECIOS
            ref_fertrac_idx_lp = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
            
            if ref_fertrac_idx_lp:
                # IMPORTANTE: Leer referencias de INVENTARIO COPIA DESPUÉS de que ya fue ordenado por TOTAL INV
                referencias_orden_maestro = read_range_as_array(ws_inv_copia, start_data_row, last_row, ref_col_idx)
                referencias_orden_maestro = [str(r).strip() if r is not None else "" for r in referencias_orden_maestro]
                
                # Crear diccionario con el orden deseado (posición ACTUAL en INVENTARIO COPIA después de ordenar)
                orden_dict = {}
                for idx, ref in enumerate(referencias_orden_maestro):
                    if ref and ref not in ("", "None", "nan"):
                        orden_dict[str(ref).strip()] = idx
                
                log(f"  Orden maestro creado con {len(orden_dict)} referencias (basado en orden FINAL de INVENTARIO COPIA)")
                
                # Determinar última fila con datos en INV LISTA PRECIOS
                last_row_lp = ws_last_row(ws_lp, ref_fertrac_idx_lp, hr_lp)
                
                # Ajustar por pivots si existen
                pivot_top_lp = ws_first_pivot_row(ws_lp)
                if pivot_top_lp and pivot_top_lp > hr_lp:
                    last_row_lp = min(last_row_lp, pivot_top_lp - 1)
                
                log(f"  Rango de datos: fila {hr_lp + 1} a {last_row_lp}")
                
                # Leer referencias actuales de INV LISTA PRECIOS
                referencias_lp = read_range_as_array(ws_lp, hr_lp + 1, last_row_lp, ref_fertrac_idx_lp)
                
                # Verificar si ya está ordenado
                referencias_lp_limpio = [str(r).strip() if r is not None else "" for r in referencias_lp]
                
                # Comparar orden actual vs orden deseado
                necesita_ordenar = False
                for i, ref in enumerate(referencias_lp_limpio):
                    if ref in orden_dict:
                        if orden_dict[ref] != i:
                            necesita_ordenar = True
                            break
                
                if not necesita_ordenar:
                    log(f"  La hoja ya está ordenada correctamente - no se requiere acción")
                else:
                    log(f"  La hoja requiere ordenamiento - aplicando...")
                    
                    # Determinar rango completo de columnas
                    used_range_lp = ws_lp.UsedRange
                    first_col_lp = used_range_lp.Column
                    last_col_lp = first_col_lp + used_range_lp.Columns.Count - 1
                    num_cols = last_col_lp - first_col_lp + 1
                    
                    # ✅ OPTIMIZACIÓN: Leer TODO el rango de datos de una sola vez
                    rng_data = ws_lp.Range(
                        ws_lp.Cells(hr_lp + 1, first_col_lp),
                        ws_lp.Cells(last_row_lp, last_col_lp)
                    )
                    datos_completos = rng_data.Value
                    
                    # Convertir a lista de listas si es necesario
                    if datos_completos is None:
                        log("  ⚠ No hay datos para ordenar")
                    else:
                        # Si es una sola fila, datos_completos es una tupla, convertir a lista de listas
                        if not isinstance(datos_completos[0], (tuple, list)):
                            datos_completos = [datos_completos]
                        
                        log(f"  Datos leídos: {len(datos_completos)} filas x {len(datos_completos[0]) if datos_completos else 0} columnas")
                        
                        # Crear lista de tuplas (orden_deseado, índice_fila, datos_fila)
                        filas_con_orden = []
                        for i, fila_data in enumerate(datos_completos):
                            ref = referencias_lp_limpio[i]
                            if ref in orden_dict:
                                orden_deseado = orden_dict[ref]
                            else:
                                orden_deseado = 999999  # Al final
                            filas_con_orden.append((orden_deseado, list(fila_data)))
                        
                        # Ordenar por el orden deseado
                        filas_con_orden.sort(key=lambda x: x[0])
                        
                        # Extraer solo los datos ordenados
                        datos_ordenados = [fila for orden, fila in filas_con_orden]
                        
                        log(f"  Escribiendo {len(datos_ordenados)} filas ordenadas...")
                        
                        # ✅ OPTIMIZACIÓN: Escribir TODO de una sola vez
                        rng_data.Value = datos_ordenados
                        
                        log(f"  ✓ Hoja INV LISTA PRECIOS ordenada según orden FINAL de INVENTARIO COPIA")
                
            else:
                log("  ⚠ No se encontró columna REFERENCIA FERTRAC en INV LISTA PRECIOS")
        else:
            log("  ⚠ No se encontró la hoja INV LISTA PRECIOS")
            
    except Exception as e:
        log(f"  ❌ ERROR al ordenar INV LISTA PRECIOS: {e}")
        import traceback
        log(traceback.format_exc())


    # LIMPIEZA FINAL: Eliminar filas con REFERENCIA FERTRAC vacía en INV LISTA PRECIOS
    log("")
    log("="*70)
    log("LIMPIEZA FINAL: ELIMINANDO FILAS CON REFERENCIA FERTRAC VACÍA")
    log("="*70)
    try:
        # Buscar la hoja INV LISTA PRECIOS
        ws_lp = None
        target_norm = _norm(SHEET_INV_LISTA)
        
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == target_norm or target_norm in _norm(sheet_name):
                ws_lp = wb.Worksheets(i)
                log(f"Hoja encontrada: '{sheet_name}'")
                break
        
        if ws_lp is None:
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name_norm = _norm(wb.Worksheets(i).Name)
                if "inv" in sheet_name_norm and "lista" in sheet_name_norm and "precio" in sheet_name_norm:
                    ws_lp = wb.Worksheets(i)
                    log(f"Hoja encontrada (por palabras clave): '{wb.Worksheets(i).Name}'")
                    break
        
        if ws_lp:
            # Obtener encabezados
            hr_lp, hdr_lp, hdrn_lp = ws_headers_smart(ws_lp, HEADER_ROW_INV_LISTA, ["REFERENCIA FERTRAC"])
            ref_fertrac_idx = hdrn_lp.get(_norm("REFERENCIA FERTRAC"))
            
            if ref_fertrac_idx:
                # Determinar última fila REAL con datos en CUALQUIER columna
                # Usar UsedRange para detectar todas las filas con datos
                pivot_top_lp = ws_first_pivot_row(ws_lp)
                
                try:
                    # Obtener el rango usado completo
                    used_range = ws_lp.UsedRange
                    last_row_used = used_range.Rows.Count
                    
                    if pivot_top_lp and pivot_top_lp > hr_lp:
                        last_row_lp = min(last_row_used, pivot_top_lp - 1)
                    else:
                        last_row_lp = last_row_used
                    
                    log(f"Rango usado completo: hasta fila {last_row_used}")
                except Exception as e:
                    log(f"  ⚠ No se pudo obtener UsedRange: {e}")
                    # Fallback al método original
                    if pivot_top_lp and pivot_top_lp > hr_lp:
                        last_row_lp = pivot_top_lp - 1
                    else:
                        last_row_lp = ws_last_row(ws_lp, ref_fertrac_idx, hr_lp)
                
                log(f"Analizando filas {hr_lp + 1} a {last_row_lp}...")
                
                # Identificar filas con REFERENCIA FERTRAC vacía
                filas_a_eliminar = []
                for row_idx in range(hr_lp + 1, last_row_lp + 1):
                    try:
                        valor_ref = ws_lp.Cells(row_idx, ref_fertrac_idx).Value
                        
                        # Verificar si está vacío de múltiples formas
                        esta_vacio = False
                        
                        if valor_ref is None:
                            esta_vacio = True
                        elif isinstance(valor_ref, str):
                            # String vacío o solo espacios
                            if not valor_ref.strip():
                                esta_vacio = True
                        elif isinstance(valor_ref, (int, float)):
                            # Si es 0 o NaN, considerar como válido (no eliminar)
                            esta_vacio = False
                        else:
                            # Cualquier otro tipo vacío
                            try:
                                if str(valor_ref).strip() in ("", "None", "#N/A", "#REF!", "#VALUE!"):
                                    esta_vacio = True
                            except:
                                pass
                        
                        if esta_vacio:
                            filas_a_eliminar.append(row_idx)
                            
                    except Exception as e:
                        # Si hay error al leer, considerar como posible fila vacía
                        log(f"  ⚠ Error al leer fila {row_idx}: {e}")
                        continue
                
                if filas_a_eliminar:
                    log(f"  Encontradas {len(filas_a_eliminar)} filas con REFERENCIA FERTRAC vacía")
                    
                    # Eliminar filas de abajo hacia arriba para mantener índices correctos
                    eliminadas = 0
                    for row_idx in reversed(filas_a_eliminar):
                        try:
                            ws_lp.Rows(row_idx).Delete()
                            eliminadas += 1
                        except Exception as e:
                            log(f"  ⚠ Error al eliminar fila {row_idx}: {e}")
                    
                    log(f"  ✓ {eliminadas} filas eliminadas en INV LISTA PRECIOS")
                else:
                    log("  ℹ No se encontraron filas con REFERENCIA FERTRAC vacía")
            else:
                log("  ⚠ Columna 'REFERENCIA FERTRAC' no encontrada")
        else:
            log("  ⚠ Hoja INV LISTA PRECIOS no encontrada")
            
    except Exception as e:
        log(f"❌ ERROR al eliminar filas con REFERENCIA FERTRAC vacía: {e}")
        import traceback
        log(traceback.format_exc())
    
    log("")



    # Eliminar hoja INVENTARIO y renombrar INVENTARIO COPIA
    log("RENOMBRANDO HOJAS: Eliminando INVENTARIO y renombrando INVENTARIO COPIA...")
    try:
        # Desactivar alertas
        excel.DisplayAlerts = False
        
        # 1. Buscar y eliminar la hoja INVENTARIO original
        sheet_inventario_eliminada = False
        for i in range(1, wb.Worksheets.Count + 1):
            try:
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_ORIG):
                    log(f"  Eliminando hoja: '{sheet_name}'")
                    wb.Worksheets(i).Delete()
                    sheet_inventario_eliminada = True
                    log(f"Hoja '{sheet_name}' eliminada")
                    break
            except Exception as e:
                log(f"  ⚠ Error al eliminar hoja INVENTARIO: {e}")
        
        if not sheet_inventario_eliminada:
            log("  ⚠ No se encontró la hoja INVENTARIO para eliminar")
        
        # 2. Renombrar INVENTARIO COPIA a INVENTARIO
        sheet_renombrada = False
        for i in range(1, wb.Worksheets.Count + 1):
            try:
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_COPIA):
                    log(f"  Renombrando hoja: '{sheet_name}' → '{SHEET_INV_ORIG}'")
                    wb.Worksheets(i).Name = SHEET_INV_ORIG
                    sheet_renombrada = True
                    log(f"Hoja renombrada a '{SHEET_INV_ORIG}'")
                    break
            except Exception as e:
                log(f"  ⚠ Error al renombrar hoja: {e}")
        
        if not sheet_renombrada:
            log("  ⚠ No se encontró la hoja INVENTARIO COPIA para renombrar")
        
        # Reactivar alertas
        excel.DisplayAlerts = True
        
        # 3. Guardar cambios
        if sheet_inventario_eliminada or sheet_renombrada:
            log("Guardando cambios en el archivo...")
            wb.Save()
            log("Cambios guardados exitosamente")
        
    except Exception as e:
        log(f"❌ ERROR al renombrar hojas: {e}")
        import traceback
        log(traceback.format_exc())
        excel.DisplayAlerts = True



    # Actualizar tablas dinámicas en RESUMEN LINEA y Hoja2

    log("Actualizando tablas dinámicas...")
    try:
        # PASO 1: Determinar el rango correcto de datos (excluyendo fila de subtotales)
        ws_inv_final = None
        for i in range(1, wb.Worksheets.Count + 1):
            sheet_name = wb.Worksheets(i).Name
            if _norm(sheet_name) == _norm(SHEET_INV_ORIG):
                ws_inv_final = wb.Worksheets(i)
                break
        
        last_data_row_for_pivot = None
        first_col = 1
        last_col = 1
        
        if ws_inv_final:
            log("  Determinando rango de datos para tablas dinámicas...")
            
            # Obtener encabezados
            hdr_final, hdrn_final = ws_headers(ws_inv_final, HEADER_ROW_INV)
            
            # Buscar columna EXISTENCIA para detectar última fila con datos
            exist_col_final = None
            for name, col in hdr_final.items():
                if str(name).upper().startswith("EXISTENCIA"):
                    exist_col_final = col
                    break
            
            if exist_col_final:
                # Encontrar última fila con datos
                last_row_temp = ws_last_row(ws_inv_final, exist_col_final, HEADER_ROW_INV)
                
                # Verificar si la última fila contiene SUBTOTAL (fórmula)
                try:
                    cell_formula = ws_inv_final.Cells(last_row_temp, exist_col_final).Formula
                    
                    if cell_formula and "SUBTOTAL" in str(cell_formula).upper():
                        # Es una fila de subtotales - excluirla del rango de la tabla dinámica
                        last_data_row_for_pivot = last_row_temp - 1
                        log(f"  ✓ Fila de subtotales detectada en fila {last_row_temp}")
                        log(f"  ✓ Rango para tabla dinámica: hasta fila {last_data_row_for_pivot}")
                    else:
                        # No hay subtotal al final
                        last_data_row_for_pivot = last_row_temp
                        log(f"  ℹ No se detectó fila de subtotales - usando fila {last_data_row_for_pivot}")
                except Exception as e:
                    last_data_row_for_pivot = last_row_temp
                    log(f"  ⚠ Error al verificar subtotales: {e}")
            
            # Determinar rango de columnas
            try:
                used_range = ws_inv_final.UsedRange
                first_col = used_range.Column
                last_col = first_col + used_range.Columns.Count - 1
            except:
                pass
        
        # PASO 2: Actualizar tablas dinámicas
        hojas_para_actualizar = ["RESUMEN LINEA", "Hoja2"]
        tablas_actualizadas = 0
        
        for nombre_hoja in hojas_para_actualizar:
            try:
                # Buscar la hoja
                ws_pivot = None
                nombre_normalizado = _norm(nombre_hoja)
                
                for i in range(1, wb.Worksheets.Count + 1):
                    sheet_name = wb.Worksheets(i).Name
                    if _norm(sheet_name) == nombre_normalizado:
                        ws_pivot = wb.Worksheets(i)
                        log(f"   Procesando hoja: '{sheet_name}'")
                        break
                
                if not ws_pivot:
                    log(f"  ⚠ Hoja '{nombre_hoja}' no encontrada")
                    continue
                
                # Obtener el número de tablas dinámicas en la hoja
                try:
                    pivot_count = int(getattr(ws_pivot.PivotTables(), "Count", 0))
                except:
                    pivot_count = 0
                
                if pivot_count == 0:
                    log(f"  ⚠ No se encontraron tablas dinámicas en '{sheet_name}'")
                    continue
                
                # Actualizar cada tabla dinámica de la hoja
                log(f"  Actualizando {pivot_count} tabla(s) dinámica(s)...")
                
                for j in range(1, pivot_count + 1):
                    try:
                        pivot_table = ws_pivot.PivotTables(j)
                        
                        # Obtener el nombre de la tabla dinámica si existe
                        try:
                            pivot_name = pivot_table.Name
                            log(f"    - Actualizando tabla: {pivot_name}")
                        except:
                            log(f"    - Actualizando tabla {j}")
                        
                        # CRÍTICO: Cambiar el rango de origen ANTES de refrescar
                        if last_data_row_for_pivot and ws_inv_final:
                            try:
                                # Construir referencia del nuevo rango (excluyendo subtotales)
                                col_letter_last = _col_num_to_letter(last_col)
                                new_source = f"INVENTARIO!$A${HEADER_ROW_INV}:${col_letter_last}${last_data_row_for_pivot}"
                                
                                # Crear nuevo PivotCache con el rango correcto
                                new_cache = wb.PivotCaches().Create(
                                    SourceType=1,  # xlDatabase
                                    SourceData=new_source
                                )
                                
                                # Cambiar el cache de la tabla dinámica
                                pivot_table.ChangePivotCache(new_cache)
                                log(f"      ✓ Rango actualizado a: {new_source}")
                                
                            except Exception as e:
                                log(f"      ⚠ No se pudo cambiar rango (se usará el existente): {e}")
                        
                        # Refrescar la tabla dinámica
                        pivot_table.RefreshTable()
                        tablas_actualizadas += 1
                        log(f"      ✓ Tabla actualizada")
                        
                    except Exception as e:
                        log(f"      ✗ Error al actualizar tabla {j}: {e}")
                        
            except Exception as e:
                log(f"  ✗ Error al procesar hoja '{nombre_hoja}': {e}")
        
        if tablas_actualizadas > 0:
            log(f"✓ {tablas_actualizadas} tabla(s) dinámica(s) actualizada(s) exitosamente")
        else:
            log("⚠ No se actualizaron tablas dinámicas")
        
    except Exception as e:
        log(f"❌ ERROR al actualizar tablas dinámicas: {e}")
        import traceback
        log(traceback.format_exc())

    # Establecer zoom al 80% en TODAS las hojas
    log("Estableciendo zoom al 80% en todas las hojas...")
    try:
        hojas_procesadas = 0
        
        for i in range(1, wb.Worksheets.Count + 1):
            try:
                ws = wb.Worksheets(i)
                sheet_name = ws.Name
                
                # Activar la hoja
                ws.Activate()
                
                # Establecer zoom
                excel.ActiveWindow.Zoom = 80
                hojas_procesadas += 1
                
            except Exception as e:
                log(f"  ⚠ Error en hoja {i}: {e}")        
       
    # NUEVO: Centrar columna EXISTENCIA en hoja INVENTARIO
        log("Centrando columna EXISTENCIA...")
        try:
            # Buscar la hoja INVENTARIO
            ws_inv_final = None
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_ORIG):
                    ws_inv_final = wb.Worksheets(i)
                    break
            
            if ws_inv_final:
                # CORRECCIÓN: Usar la fila correcta de encabezados (fila 2 en el archivo)
                hdr_final, hdrn_final = ws_headers(ws_inv_final, HEADER_ROW_INV)
                
                # Buscar columna EXISTENCIA usando el diccionario ORIGINAL
                exist_col_final = None
                for name, col in hdr_final.items():
                    name_upper = str(name).upper()
                    if name_upper.startswith("EXISTENCIA"):
                        exist_col_final = col
                        log(f"  Columna encontrada: '{name}' (índice {col})")
                        break
                
                if exist_col_final:
                    # Centrar toda la columna EXISTENCIA
                    ws_inv_final.Columns(exist_col_final).HorizontalAlignment = -4108  # xlCenter

                else:
                    log("  ⚠ Columna EXISTENCIA no encontrada para centrar")
                    log(f"  Encabezados encontrados: {list(hdr_final.keys())[:15]}")
            else:
                log(f"  ⚠ No se encontró la hoja '{SHEET_INV_ORIG}' para centrar EXISTENCIA")
                
        except Exception as e:
            log(f"  ⚠ Error al centrar columna EXISTENCIA: {e}")
            import traceback
            log(traceback.format_exc())
        
        # Activar la hoja INVENTARIO al final
        log("Activando hoja INVENTARIO como hoja predeterminada...")
        try:
            ws_inventario_final = None
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_ORIG):
                    ws_inventario_final = wb.Worksheets(i)
                    break
            
            if ws_inventario_final:
                ws_inventario_final.Activate()
                # Asegurar que la celda A1 esté seleccionada
                ws_inventario_final.Range("A1").Select()

            else:
                log(f"  ⚠ No se encontró la hoja '{SHEET_INV_ORIG}'")
        except Exception as e:
            log(f"  ⚠ Error al activar hoja INVENTARIO: {e}")
           # ===== ACTIVAR FILTROS EN FILA 2 =====
        log("Activando filtros en fila 2 de INVENTARIO...")
        try:
            # Buscar la hoja INVENTARIO
            ws_inv_final = None
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_ORIG):
                    ws_inv_final = wb.Worksheets(i)
                    break
            
            if ws_inv_final:
                # Activar la hoja
                ws_inv_final.Activate()
                
                # Obtener el rango usado
                used_range = ws_inv_final.UsedRange
                last_col = used_range.Columns.Count
                
                # Determinar última fila (considerando pivots)
                pivot_top = ws_first_pivot_row(ws_inv_final)
                if pivot_top and pivot_top > HEADER_ROW_INV:
                    last_row = pivot_top - 1
                else:
                    last_row = used_range.Rows.Count
                
                # Desactivar filtro si ya existe
                if ws_inv_final.AutoFilterMode:
                    ws_inv_final.AutoFilterMode = False
                
                # Definir el rango para el filtro (desde fila 2 hasta la última fila)
                filter_range = ws_inv_final.Range(
                    ws_inv_final.Cells(HEADER_ROW_INV, 1),
                    ws_inv_final.Cells(last_row, last_col)
                )
                
                # Activar AutoFilter
                filter_range.AutoFilter(Field=1)
                
                log(f"✅ Filtros activados en fila {HEADER_ROW_INV}")
                
            else:
                log(f"  ⚠️ No se encontró la hoja '{SHEET_INV_ORIG}'")
                
        except Exception as e:
            log(f"  ⚠️ Error al activar filtros: {e}")
            import traceback
            log(traceback.format_exc())    
        
        # ===== ESTABLECER ALTO DE FILAS =====
        log("Estableciendo alto de filas en 14,5...")
        try:
            # Buscar la hoja INVENTARIO
            ws_inv_final = None
            for i in range(1, wb.Worksheets.Count + 1):
                sheet_name = wb.Worksheets(i).Name
                if _norm(sheet_name) == _norm(SHEET_INV_ORIG):
                    ws_inv_final = wb.Worksheets(i)
                    break
            
            if ws_inv_final:
                # Determinar última fila con datos (considerando pivots)
                pivot_top = ws_first_pivot_row(ws_inv_final)
                if pivot_top and pivot_top > HEADER_ROW_INV:
                    ultima_fila = pivot_top - 1
                else:
                    # Usar el rango usado para determinar la última fila
                    used_range = ws_inv_final.UsedRange
                    ultima_fila = used_range.Rows.Count
                
                # Establecer alto de 14.5 para todas las filas desde la 3 hasta la última
                if ultima_fila > 2:
                    rango_filas = f"3:{ultima_fila}"
                    ws_inv_final.Rows(rango_filas).RowHeight = 14.5
                    log(f"✅ Alto de fila establecido en 14,5 para filas 3 a {ultima_fila}")
                else:
                    log("  ℹ No hay filas después de la fila 2 para modificar")
                
            else:
                log(f"  ⚠️ No se encontró la hoja '{SHEET_INV_ORIG}'")
                
        except Exception as e:
            log(f"  ⚠️ Error al establecer alto de filas: {e}")
            import traceback
            log(traceback.format_exc())

        # GUARDAR después de establecer el zoom y activar la hoja
        log("Guardando cambios con zoom aplicado y hoja INVENTARIO activa...")
        wb.Save()
        log("✅ Cambios guardados exitosamente")
        
    except Exception as e:
        log(f"❌ ERROR al establecer zoom: {e}")
        import traceback
        log(traceback.format_exc())

    # Cerrar MATRIZ USD
    if matriz_wb:
        try:
            log("")
            log("Cerrando MATRIZ USD...")
            matriz_wb.Close(SaveChanges=False)
            log("  ✓ MATRIZ USD cerrado")
        except Exception as e:
            log(f"  ⚠ Error al cerrar MATRIZ USD: {e}")

    # Limpiar archivo temporal
    if matriz_tmp_path and os.path.exists(matriz_tmp_path):
        try:
            os.remove(matriz_tmp_path)
            log("  ✓ Archivo temporal eliminado")
        except Exception as e:
            log(f"  ⚠ No se pudo eliminar temporal: {e}")


    # Cerrar Excel
    excel_close(excel, wb, save=False)

    tmp = saveinfo.get("tmp_path")
    if tmp and os.path.exists(tmp):
        with contextlib.suppress(Exception):
            os.remove(tmp)

    log("== Proceso completado exitosamente ==")


if __name__ == "__main__":
    main()